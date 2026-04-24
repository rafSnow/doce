import customtkinter as ctk
import os
from pathlib import Path
from tkinter import ttk, messagebox, filedialog
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.formatters import fmt_data, fmt_moeda, parse_data, parse_float
from app.models.pedido import Pedido
from app.models.pedido_item import PedidoItem
from app.services.configuracao_service import ConfiguracaoService
from app.services.cliente_service import ClienteService
from app.services.insumo_service import InsumoService
from app.services.pedido_service import PedidoService
from app.services.produto_service import ProdutoService
from app.services.recibo_service import ReciboService
from app.core.enums import StatusPagamento, FormaPagamento
from app.ui.theme import (
    ACCENT,
    BG_DEEP,
    CARD_BG,
    CARD_BORDER,
    COLOR_GREEN,
    FIELD_BG,
    HEADER_BG,
    TEXT_MUTED,
    TEXT_PRIMARY,
    TEXT_SECONDARY,
    _combo,
    _entry,
    _optmenu,
    _treeview_style,
)


class PedidosView(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=BG_DEEP, **kwargs)

        self.pedido_service  = PedidoService()
        self.cliente_service = ClienteService()
        self.produto_service = ProdutoService()
        self.recibo_service  = ReciboService()
        self.config_service  = ConfiguracaoService()
        self.insumo_service  = InsumoService()
        self.current_pedido_id = None
        self.current_itens: List[PedidoItem] = []

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ── Header ────────────────────────────────────────────────────────
        header_frame = ctk.CTkFrame(self, fg_color=HEADER_BG, corner_radius=0, height=56)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(10, 0))
        header_frame.grid_propagate(False)
        header_frame.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            header_frame,
            text="Gerenciamento de Pedidos",
            text_color=TEXT_PRIMARY,
            font=ctk.CTkFont(family="Roboto", size=20, weight="bold"),
        )
        self.title_label.grid(row=0, column=0, padx=20, sticky="w")

        actions = ctk.CTkFrame(header_frame, fg_color="transparent")
        actions.grid(row=0, column=1, padx=(6, 16), sticky="e")

        self.btn_novo = ctk.CTkButton(
            actions, text="+ Novo Pedido", command=self._novo_pedido,
            fg_color=ACCENT, hover_color="#A84F5E", text_color="#FFFFFF",
            height=30, corner_radius=20, font=ctk.CTkFont(size=12, weight="bold"),
        )
        self.btn_novo.grid(row=0, column=0, padx=4)

        for col, text, cmd in [
            (1, "Pré-visualizar", self._previsualizar_recibo),
            (2, "Salvar PDF",     self._salvar_recibo_pdf),
            (3, "Imprimir",       self._imprimir_recibo),
            (4, "Exportar Excel", self._exportar_excel),
        ]:
            ctk.CTkButton(
                actions, text=text, command=cmd,
                fg_color=FIELD_BG, hover_color="#F2D5DC",
                border_color=CARD_BORDER, border_width=1,
                text_color=TEXT_SECONDARY, height=30, corner_radius=20,
                font=ctk.CTkFont(size=12),
            ).grid(row=0, column=col, padx=4)

        # guarda referências para os botões de recibo (usados em _show_lista/_show_form)
        self.btn_previsualizar_recibo = actions.grid_slaves(row=0, column=1)[0]
        self.btn_salvar_recibo        = actions.grid_slaves(row=0, column=2)[0]
        self.btn_imprimir_recibo      = actions.grid_slaves(row=0, column=3)[0]
        self.btn_exportar             = actions.grid_slaves(row=0, column=4)[0]

        ctk.CTkFrame(self, fg_color=CARD_BORDER, height=1).grid(row=1, column=0, sticky="ew")

        # ── Body ──────────────────────────────────────────────────────────
        self.body_container = ctk.CTkFrame(self, fg_color="transparent")
        self.body_container.grid(row=2, column=0, padx=14, pady=(10, 12), sticky="nsew")
        self.body_container.grid_columnconfigure(0, weight=1)
        self.body_container.grid_rowconfigure(0, weight=1)

        self.view_lista = ctk.CTkFrame(self.body_container, fg_color="transparent")
        self.view_form  = ctk.CTkFrame(self.body_container, fg_color="transparent")

        self._setup_lista_view()
        self._setup_form_view()

        self._show_lista()
        self._carregar_clientes()
        self._carregar_pedidos()

    def refresh(self):
        self._carregar_clientes()
        self._carregar_pedidos()

    # ── visibilidade ──────────────────────────────────────────────────────
    def _show_lista(self):
        self.view_form.grid_forget()
        self.view_lista.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="normal")
        
        self.btn_previsualizar_recibo.configure(state="disabled")
        self.btn_salvar_recibo.configure(state="disabled")
        self.btn_imprimir_recibo.configure(state="disabled")
        
        self.title_label.configure(text="Gerenciamento de Pedidos")

    def _show_form(self, editando=False):
        self.view_lista.grid_forget()
        self.view_form.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="disabled")
        
        self.btn_previsualizar_recibo.configure(state="normal")
        self.btn_salvar_recibo.configure(state="normal")
        self.btn_imprimir_recibo.configure(state="normal")
        
        self.title_label.configure(text="Editar Pedido" if editando else "Novo Pedido")
        self.cb_cliente.focus_set()

    # ── máscaras ──────────────────────────────────────────────────────────
    def _aplicar_mascara_data(self, event):
        entry  = event.widget
        digits = "".join(ch for ch in entry.get() if ch.isdigit())[:8]
        partes = []
        if len(digits) >= 2: partes.append(digits[:2])
        elif digits:          partes.append(digits)
        if len(digits) >= 4:  partes.append(digits[2:4])
        elif len(digits) > 2: partes.append(digits[2:])
        if len(digits) > 4:   partes.append(digits[4:])
        entry.delete(0, "end"); entry.insert(0, "/".join(partes))

    def _aplicar_mascara_decimal(self, event):
        entry = event.widget
        bruto = entry.get()
        digitos    = "".join(ch for ch in bruto if ch.isdigit())
        tem_virgula = "," in bruto
        if not digitos:
            sanitizado = ""
        else:
            if tem_virgula:
                parte_int, parte_dec = bruto.split(",", 1)
                int_digits = "".join(ch for ch in parte_int if ch.isdigit()) or "0"
                dec_all    = "".join(ch for ch in parte_dec if ch.isdigit())
                if len(dec_all) > 2:
                    int_digits = int_digits + dec_all[:-2]; dec_digits = dec_all[-2:]
                else:
                    dec_digits = dec_all.ljust(2, "0")
            else:
                int_digits = digitos; dec_digits = "00"
            int_norm = str(int(int_digits))
            grupos = []
            while int_norm:
                grupos.insert(0, n[-3:]); int_norm = int_norm[:-3]
            sanitizado = f"{'.'.join(grupos)},{dec_digits}"
        entry.delete(0, "end"); entry.insert(0, sanitizado)

    def _aplicar_mascara_inteiro(self, event):
        entry = event.widget
        valor = entry.get()
        digits = "".join(ch for ch in valor if ch.isdigit())
        entry.delete(0, "end")
        entry.insert(0, digits)

    # ── recibo ────────────────────────────────────────────────────────────
    def _carregar_clientes(self):
        clientes = self.cliente_service.listar()
        valores  = [c.nome for c in clientes if (c.nome or "").strip()]
        self.cb_cliente.configure(values=valores or [""])
        if valores and not self.cb_cliente.get().strip():
            self.cb_cliente.set(valores[0])

    def _coletar_pedido_formulario(self) -> Pedido:
        nome_cliente = self.cb_cliente.get().strip()
        if not nome_cliente:    raise ValueError("Informe o nome do cliente.")
        if not self.current_itens: raise ValueError("Adicione ao menos um item ao pedido.")

        data_pedido      = parse_data(self.entry_data_pedido.get(), "Data do Pedido", obrigatorio=True)
        data_entrega     = parse_data(self.entry_data_entrega.get(), "Data de Entrega")
        pag_inicial_data = parse_data(self.entry_pag_ini_data.get(), "Data do Pagamento Inicial")
        pag_final_data   = parse_data(self.entry_pag_fin_data.get(), "Data do Pagamento Final")

        if data_entrega and datetime.strptime(data_entrega, "%d/%m/%Y") < datetime.strptime(data_pedido, "%d/%m/%Y"):
            raise ValueError("A Data de Entrega não pode ser menor que a Data do Pedido.")

        responsavel = self.entry_responsavel.get().strip()
        if not responsavel: raise ValueError("O campo Responsável é obrigatório.")

        pag_inicial_valor = parse_float(self.entry_pag_ini_valor.get(), "Valor do Pagamento Inicial", minimo=0.0)
        pag_final_valor   = parse_float(self.entry_pag_fin_valor.get(), "Valor do Pagamento Final", minimo=0.0)

        for v in (self.cb_pag_ini_forma.get(), self.cb_pag_ini_status.get(),
                  self.cb_pag_fin_forma.get(), self.cb_pag_fin_status.get()):
            if not v.strip(): raise ValueError("Preencha forma e status dos pagamentos.")

        pedido = Pedido(
            id=self.current_pedido_id,
            cliente_nome=nome_cliente,
            data_pedido=data_pedido, data_entrega=data_entrega,
            valor_total=0.0,
            pag_inicial_valor=pag_inicial_valor, pag_inicial_data=pag_inicial_data,
            pag_inicial_forma=self.cb_pag_ini_forma.get(), pag_inicial_status=self.cb_pag_ini_status.get(),
            pag_final_valor=pag_final_valor,   pag_final_data=pag_final_data,
            pag_final_forma=self.cb_pag_fin_forma.get(),   pag_final_status=self.cb_pag_fin_status.get(),
            responsavel=responsavel,
        )
        pedido.itens = list(self.current_itens)
        return pedido

    def _pedido_para_recibo(self) -> Pedido:
        pedido = self._coletar_pedido_formulario()
        pedido.id = self.current_pedido_id
        return pedido

    def _previsualizar_recibo(self):
        try:
            caminho_tmp = self.recibo_service.gerar_previsao_recibo(
                self._pedido_para_recibo(),
                nome_estabelecimento=self.config_service.get_nome_estabelecimento(),
            )
            self.recibo_service.abrir_pdf(caminho_tmp)
            messagebox.showinfo("Pré-visualização", "O recibo foi gerado e aberto no visualizador padrão.")
        except Exception as exc:
            messagebox.showerror("Erro", f"Não foi possível gerar a pré-visualização.\n\n{exc}")

    def _salvar_recibo_pdf(self):
        try:
            pedido = self._pedido_para_recibo()
        except ValueError as exc:
            messagebox.showerror("Erro", str(exc)); return
        caminho = filedialog.asksaveasfilename(
            title="Salvar recibo em PDF", defaultextension=".pdf",
            filetypes=[("Documento PDF", "*.pdf")],
            initialfile=f"recibo_pedido_{pedido.id if pedido.id is not None else 'previsao'}.pdf",
        )
        if not caminho: return
        try:
            self.recibo_service.gerar_recibo_pdf(
                pedido, caminho,
                nome_estabelecimento=self.config_service.get_nome_estabelecimento(),
            )
            messagebox.showinfo("Sucesso", f"Recibo salvo com sucesso em:\n{caminho}")
        except Exception as exc:
            messagebox.showerror("Erro", f"Não foi possível salvar o recibo.\n\n{exc}")

    def _imprimir_recibo(self):
        try:
            pedido = self._pedido_para_recibo()
            caminho_tmp = self.recibo_service.gerar_previsao_recibo(
                pedido,
                nome_estabelecimento=self.config_service.get_nome_estabelecimento(),
            )
            self.recibo_service.imprimir_pdf(caminho_tmp)
            messagebox.showinfo("Impressão", "O recibo foi enviado para a impressora padrão.")
        except Exception as exc:
            messagebox.showerror("Erro", f"Não foi possível imprimir o recibo.\n\n{exc}")

    # ── treeview style ────────────────────────────────────────────────────
    def _apply_tree_style(self):
        _treeview_style("Pedidos", rowheight=27)

    # ── lista ─────────────────────────────────────────────────────────────
    def _setup_lista_view(self):
        self._apply_tree_style()
        self.view_lista.grid_columnconfigure(0, weight=1)
        self.view_lista.grid_rowconfigure(1, weight=1)

        # filtros
        filter_frame = ctk.CTkFrame(self.view_lista, fg_color=CARD_BG,
                                    corner_radius=12, border_width=1, border_color=CARD_BORDER)
        filter_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(filter_frame, text="Filtros", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).pack(side="left", padx=(12, 6), pady=10)

        ctk.CTkLabel(filter_frame, text="Cliente:", text_color=TEXT_SECONDARY).pack(side="left", padx=(8, 4))
        self.filtro_cliente = _entry(filter_frame, width=260)
        self.filtro_cliente.pack(side="left", padx=5)

        ctk.CTkLabel(filter_frame, text="Status Pag.:", text_color=TEXT_SECONDARY).pack(side="left", padx=(10, 4))
        self.filtro_status = _optmenu(filter_frame, ["Todos", StatusPagamento.PENDENTE.value, StatusPagamento.RECEBIDO.value])
        self.filtro_status.pack(side="left", padx=5)
        self.filtro_status.set("Todos")

        ctk.CTkButton(filter_frame, text="Buscar", command=self._carregar_pedidos,
                      width=100, fg_color=ACCENT, hover_color="#A84F5E",
                      text_color="#FFFFFF", corner_radius=10,
                      ).pack(side="left", padx=10)

        # tabela
        tree_frame = ctk.CTkFrame(self.view_lista, fg_color=CARD_BG,
                                  corner_radius=12, border_width=1, border_color=CARD_BORDER)
        tree_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        columns = ("id", "cliente", "data", "entrega", "valor_total", "lucro_total", "responsavel")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", style="Pedidos.Treeview")
        for col, txt, w, anc in [
            ("id",          "ID",            50,  "center"),
            ("cliente",     "Cliente",       200, "w"),
            ("data",        "Data Pedido",   100, "center"),
            ("entrega",     "Data Entrega",  100, "center"),
            ("valor_total", "Valor Total",   100, "e"),
            ("lucro_total", "Lucro Total",   100, "e"),
            ("responsavel", "Responsável",   150, "w"),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor=anc)
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.bind("<Double-1>", self._editar_pedido)

    # ── formulário ────────────────────────────────────────────────────────
    def _setup_form_view(self):
        self.view_form.grid_columnconfigure((0, 1), weight=1)

        # Atalhos de teclado
        self.view_form.bind("<Control-s>", lambda e: self._salvar())
        self.view_form.bind("<Escape>",    lambda e: self._cancelar())

        # Dados Gerais
        fg = ctk.CTkFrame(self.view_form, fg_color=CARD_BG, corner_radius=12,
                          border_width=1, border_color=CARD_BORDER)
        fg.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(fg, text="Dados Gerais", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, columnspan=4, padx=12, pady=(10, 8), sticky="w")

        ctk.CTkLabel(fg, text="Nome do Cliente*:", text_color=TEXT_SECONDARY).grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.cb_cliente = _combo(fg, values=[], width=300)
        self.cb_cliente.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(fg, text="Digite um nome novo ou selecione um cliente já cadastrado.",
                     text_color=TEXT_MUTED, font=ctk.CTkFont(size=10),
                     ).grid(row=3, column=0, columnspan=4, padx=5, pady=(0, 6), sticky="w")

        ctk.CTkLabel(fg, text="Data Pedido*:", text_color=TEXT_SECONDARY).grid(row=1, column=2, padx=5, pady=5, sticky="e")
        self.entry_data_pedido = _entry(fg, width=150, placeholder_text="DD/MM/AAAA")
        self.entry_data_pedido.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.entry_data_pedido.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(fg, text="Data Entrega:", text_color=TEXT_SECONDARY).grid(row=2, column=2, padx=5, pady=5, sticky="e")
        self.entry_data_entrega = _entry(fg, width=150, placeholder_text="DD/MM/AAAA")
        self.entry_data_entrega.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.entry_data_entrega.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(fg, text="Responsável*:", text_color=TEXT_SECONDARY).grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.entry_responsavel = _entry(fg, width=200, placeholder_text="Texto")
        self.entry_responsavel.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(fg, text="* Campos obrigatórios", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10)).grid(row=4, column=0, columnspan=4, padx=12, pady=(0, 5), sticky="w")

        # Pagamentos
        fp = ctk.CTkFrame(self.view_form, fg_color=CARD_BG, corner_radius=12,
                          border_width=1, border_color=CARD_BORDER)
        fp.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        for col_offset, titulo, prefix in [
            (0, "SINAL (Pagamento Inicial)",   "ini"),
            (4, "RESTANTE (Pagamento Final)",   "fin"),
        ]:
            ctk.CTkLabel(fp, text=titulo, font=ctk.CTkFont(size=14, weight="bold"),
                         text_color=TEXT_PRIMARY
                         ).grid(row=0, column=col_offset, columnspan=4, pady=(10, 5))

            ctk.CTkLabel(fp, text="Valor R$:", text_color=TEXT_SECONDARY
                         ).grid(row=1, column=col_offset, padx=5, pady=5, sticky="e")
            ev = _entry(fp, width=100, placeholder_text="0,00")
            ev.grid(row=1, column=col_offset+1, padx=5, pady=5, sticky="w")
            setattr(self, f"entry_pag_{prefix}_valor", ev)

            ctk.CTkLabel(fp, text="Data:", text_color=TEXT_SECONDARY
                         ).grid(row=1, column=col_offset+2, padx=5, pady=5, sticky="e")
            ed = _entry(fp, width=100, placeholder_text="DD/MM/AAAA")
            ed.grid(row=1, column=col_offset+3, padx=5, pady=5, sticky="w")
            ed.bind("<KeyRelease>", self._aplicar_mascara_data)
            setattr(self, f"entry_pag_{prefix}_data", ed)

            ctk.CTkLabel(fp, text="Forma:", text_color=TEXT_SECONDARY
                         ).grid(row=2, column=col_offset, padx=5, pady=5, sticky="e")
            ef = _combo(fp, [FormaPagamento.DINHEIRO.value, FormaPagamento.PIX.value, FormaPagamento.CARTAO_CREDITO.value, FormaPagamento.CARTAO_DEBITO.value, FormaPagamento.BOLETO.value, FormaPagamento.TRANSFERENCIA.value], width=150)
            ef.grid(row=2, column=col_offset+1, padx=5, pady=5, sticky="w")
            setattr(self, f"cb_pag_{prefix}_forma", ef)

            ctk.CTkLabel(fp, text="Status:", text_color=TEXT_SECONDARY
                         ).grid(row=2, column=col_offset+2, padx=5, pady=5, sticky="e")
            es = _combo(fp, [StatusPagamento.PENDENTE.value, StatusPagamento.RECEBIDO.value], width=100)
            es.grid(row=2, column=col_offset+3, padx=5, pady=5, sticky="w")
            setattr(self, f"cb_pag_{prefix}_status", es)

        # Itens
        frame_itens = ctk.CTkFrame(self.view_form, fg_color=CARD_BG, corner_radius=12,
                                   border_width=1, border_color=CARD_BORDER)
        frame_itens.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.view_form.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(frame_itens, text="Itens do Pedido",
                     font=ctk.CTkFont(size=14, weight="bold"), text_color=TEXT_PRIMARY
                     ).pack(anchor="w", padx=5, pady=5)

        add_item_frame = ctk.CTkFrame(frame_itens, fg_color="transparent")
        add_item_frame.pack(fill="x", padx=5, pady=5)

        ctk.CTkLabel(add_item_frame, text="Produto:", text_color=TEXT_SECONDARY).pack(side="left", padx=5)
        self.cb_add_produto = _combo(add_item_frame, values=[], width=250)
        self.cb_add_produto.pack(side="left", padx=5)

        ctk.CTkLabel(add_item_frame, text="Qtd:", text_color=TEXT_SECONDARY).pack(side="left", padx=5)
        self.entry_add_qtd = _entry(add_item_frame, width=60, placeholder_text="0")
        self.entry_add_qtd.pack(side="left", padx=5)
        self.entry_add_qtd.bind("<KeyRelease>", self._aplicar_mascara_inteiro)

        ctk.CTkButton(add_item_frame, text="Adicionar Item", command=self._add_item,
                      width=120, fg_color=ACCENT, hover_color="#A84F5E",
                      text_color="#FFFFFF", corner_radius=10,
                      ).pack(side="left", padx=10)

        ctk.CTkButton(add_item_frame, text="Remover Selecionado", command=self._remover_item,
                      width=150, fg_color="#FAE8EC", hover_color="#F2D5DC",
                      border_color=CARD_BORDER, border_width=1,
                      text_color=ACCENT, corner_radius=10,
                      ).pack(side="left", padx=10)

        self.tree_itens = ttk.Treeview(
            frame_itens,
            columns=("id","produto","qtd","preco","subtotal"),
            show="headings", height=5, style="Pedidos.Treeview",
        )
        for col, txt, w, anc in [
            ("id",       "ID Prod",              50,  "center"),
            ("produto",  "Produto",             250,  "w"),
            ("qtd",      "Quantidade",           80,  "center"),
            ("preco",    "Preço Unit. (Snapshot)", 120, "e"),
            ("subtotal", "Subtotal",            120,  "e"),
        ]:
            self.tree_itens.heading(col, text=txt)
            self.tree_itens.column(col, width=w, anchor=anc)
        self.tree_itens.pack(fill="both", expand=True, padx=5, pady=5)

        self.lbl_total = ctk.CTkLabel(
            frame_itens, text="Total: R$ 0,00",
            font=ctk.CTkFont(size=16, weight="bold"), text_color=ACCENT,
        )
        self.lbl_total.pack(side="right", padx=10, pady=10)

        self.lbl_lucro_total = ctk.CTkLabel(
            frame_itens, text="Lucro Total: R$ 0,00",
            font=ctk.CTkFont(size=14, weight="bold"), text_color=COLOR_GREEN,
        )
        self.lbl_lucro_total.pack(side="right", padx=16, pady=10)

        # Ações
        frame_acoes = ctk.CTkFrame(self.view_form, fg_color="transparent")
        frame_acoes.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky="e")

        self.btn_salvar = ctk.CTkButton(
            frame_acoes, text="Salvar", command=self._salvar,
            fg_color=ACCENT, hover_color="#A84F5E", text_color="#FFFFFF", corner_radius=10,
        )
        self.btn_salvar.pack(side="right", padx=5)

        self.btn_excluir = ctk.CTkButton(
            frame_acoes, text="Excluir",
            fg_color="#FAE8EC", hover_color="#F2D5DC",
            border_color=CARD_BORDER, border_width=1,
            text_color=ACCENT, corner_radius=10, command=self._excluir,
        )
        self.btn_excluir.pack(side="right", padx=5)

        self.btn_cancelar = ctk.CTkButton(
            frame_acoes, text="Cancelar",
            fg_color=FIELD_BG, hover_color="#F2D5DC",
            border_color=CARD_BORDER, border_width=1,
            text_color=TEXT_SECONDARY, corner_radius=10, command=self._cancelar,
        )
        self.btn_cancelar.pack(side="right", padx=5)

        self._carregar_combos()

    # ── dados ─────────────────────────────────────────────────────────────
    def _carregar_combos(self):
        produtos = self.produto_service.listar()
        self.cb_add_produto.configure(values=[f"{p.id} - {p.nome}" for p in produtos])

    def _carregar_pedidos(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        pedidos = self.pedido_service.listar(
            cliente_nome=self.filtro_cliente.get(),
            status_pagamento=self.filtro_status.get(),
        )
        for p in pedidos:
            lucro = self.pedido_service.calcular_lucro_pedido(int(p.id)) if p.id is not None else 0.0
            self.tree.insert("", "end", values=(
                p.id, p.cliente_nome, p.data_pedido, p.data_entrega,
                f"R$ {fmt_moeda(p.valor_total)}",
                f"R$ {fmt_moeda(lucro)}",
                p.responsavel,
            ))

    def _exportar_excel(self):
        pedidos = self.pedido_service.listar(
            cliente_nome=self.filtro_cliente.get().strip(),
            status_pagamento=self.filtro_status.get().strip() or "Todos",
        )
        if not pedidos:
            messagebox.showinfo("Exportar Excel", "Não há pedidos para exportar."); return
        caminho = filedialog.asksaveasfilename(
            title="Salvar exportação de pedidos", defaultextension=".xlsx",
            filetypes=[("Planilha Excel","*.xlsx")],
            initialfile=f"pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not caminho: return
        try:
            wb = Workbook()
            ws_res = wb.active; ws_res.title = "Pedidos"
            cab_res = ["ID","Cliente","Data Pedido","Data Entrega","Valor Total (R$)",
                       "Pag. Inicial","Status Inicial","Pag. Final","Status Final","Responsável"]
            ws_res.append(cab_res)
            for c in ws_res[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C96B7A")
                c.alignment = Alignment(horizontal="center", vertical="center")

            ws_it = wb.create_sheet("Itens Pedido")
            cab_it = ["Pedido ID","Cliente","Produto ID","Produto","Quantidade",
                      "Preço Unit. Snapshot (R$)","Valor Item (R$)","Data Pedido",
                      "Status Inicial","Status Final"]
            ws_it.append(cab_it)
            for c in ws_it[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="7A4A55")
                c.alignment = Alignment(horizontal="center", vertical="center")

            for p in pedidos:
                ws_res.append([p.id, p.cliente_nome,
                               fmt_data(p.data_pedido or ""),
                               fmt_data(p.data_entrega or ""),
                               p.valor_total, p.pag_inicial_valor, p.pag_inicial_status,
                               p.pag_final_valor, p.pag_final_status, p.responsavel])
                pd = self.pedido_service.get_by_id(p.id)
                if pd:
                    for item in pd.itens:
                        ws_it.append([pd.id, pd.cliente_nome, item.produto_id,
                                      item.produto_nome or "", item.quantidade,
                                      item.preco_unitario_snapshot, item.valor_item,
                                      fmt_data(pd.data_pedido or ""),
                                      pd.pag_inicial_status, pd.pag_final_status])

            for planilha in (ws_res, ws_it):
                for col in planilha.columns:
                    planilha.column_dimensions[col[0].column_letter].width = min(
                        max(len(str(c.value or "")) for c in col) + 2, 40)
                planilha.freeze_panes = "A2"

            wb.save(caminho)
            messagebox.showinfo("Exportar Excel", f"Exportação concluída.\nArquivo salvo em:\n{caminho}")
        except Exception as exc:
            messagebox.showerror("Exportar Excel", f"Não foi possível exportar.\n\n{exc}")

    def _atualizar_lista_itens(self):
        for item in self.tree_itens.get_children():
            self.tree_itens.delete(item)
        total = 0.0; lucro_total = 0.0
        for i, item in enumerate(self.current_itens):
            produto = self.produto_service.get_by_id(int(item.produto_id))
            custo   = float(produto.custo_unitario if produto else 0.0)
            self.tree_itens.insert("", "end", iid=str(i), values=(
                item.produto_id, item.produto_nome, item.quantidade,
                f"R$ {fmt_moeda(item.preco_unitario_snapshot)}",
                f"R$ {fmt_moeda(item.valor_item)}",
            ))
            total       += item.valor_item
            lucro_total += (float(item.preco_unitario_snapshot or 0) - custo) * float(item.quantidade or 0)
        self.lbl_total.configure(text=f"Total: R$ {fmt_moeda(total)}")
        self.lbl_lucro_total.configure(text=f"Lucro Total: R$ {fmt_moeda(lucro_total)}")

    def _add_item(self):
        prod_str = self.cb_add_produto.get()
        qtd_str  = self.entry_add_qtd.get().strip()
        if not prod_str or not qtd_str:
            messagebox.showwarning("Aviso", "Selecione um produto e informe a quantidade."); return
        try:
            prod_id = int(prod_str.split(" - ")[0])
            produto = self.produto_service.get_by_id(prod_id)
            if not produto: return
            qtd = int(qtd_str)
            if qtd <= 0:
                messagebox.showwarning("Aviso", "A quantidade deve ser maior que zero."); return
            item = PedidoItem(
                produto_id=produto.id, quantidade=qtd,
                preco_unitario_snapshot=produto.preco_venda_unitario,
                valor_item=produto.preco_venda_unitario * qtd,
                produto_nome=produto.nome,
            )
            existente = next((i for i in self.current_itens if i.produto_id == item.produto_id), None)
            if existente:
                existente.quantidade  += qtd
                existente.valor_item   = existente.quantidade * existente.preco_unitario_snapshot
            else:
                self.current_itens.append(item)
            self.entry_add_qtd.delete(0, "end")
            self._atualizar_lista_itens()
        except ValueError:
            messagebox.showwarning("Aviso", "Quantidade inválida.")

    def _remover_item(self):
        selected = self.tree_itens.selection()
        if not selected:
            messagebox.showwarning("Aviso", "Selecione um item para remover."); return
        del self.current_itens[int(selected[0])]
        self._atualizar_lista_itens()

    def _novo_pedido(self):
        self.current_pedido_id = None
        self.current_itens     = []
        self.cb_cliente.set("")
        self.entry_data_pedido.delete(0, "end")
        self.entry_data_pedido.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.entry_data_entrega.delete(0, "end")
        
        from app.services.configuracao_service import ConfiguracaoService
        responsavel_padrao = ConfiguracaoService().get_responsavel_padrao()
        self.entry_responsavel.delete(0, "end")
        self.entry_responsavel.insert(0, responsavel_padrao)
        
        for prefix in ("ini", "fin"):
            getattr(self, f"entry_pag_{prefix}_valor").delete(0,"end")
            getattr(self, f"entry_pag_{prefix}_valor").insert(0,"0,00")
            getattr(self, f"entry_pag_{prefix}_data").delete(0,"end")
            getattr(self, f"cb_pag_{prefix}_forma").set(FormaPagamento.PIX.value)
            getattr(self, f"cb_pag_{prefix}_status").set(StatusPagamento.PENDENTE.value)
        self.btn_excluir.configure(state="disabled")
        self._atualizar_lista_itens()
        self._show_form(editando=False)

    def _editar_pedido(self, event):
        selected = self.tree.selection()
        if not selected: return
        pedido = self.pedido_service.get_by_id(int(self.tree.item(selected[0])["values"][0]))
        if not pedido: return
        self._preencher_formulario_pedido(pedido)
        self._show_form(editando=True)

    def _preencher_formulario_pedido(self, pedido: Pedido):
        self.current_pedido_id = pedido.id
        self.current_itens = list(pedido.itens or [])

        self.cb_cliente.set(pedido.cliente_nome or "")
        self.entry_data_pedido.delete(0, "end")
        self.entry_data_pedido.insert(0, fmt_data(pedido.data_pedido or ""))
        self.entry_data_entrega.delete(0, "end")
        self.entry_data_entrega.insert(0, fmt_data(pedido.data_entrega or ""))

        self.entry_responsavel.delete(0, "end")
        self.entry_responsavel.insert(0, pedido.responsavel or "")

        for prefix, valor, data, forma, status in [
            ("ini", pedido.pag_inicial_valor, pedido.pag_inicial_data,
             pedido.pag_inicial_forma, pedido.pag_inicial_status),
            ("fin", pedido.pag_final_valor, pedido.pag_final_data,
             pedido.pag_final_forma, pedido.pag_final_status),
        ]:
            getattr(self, f"entry_pag_{prefix}_valor").delete(0, "end")
            getattr(self, f"entry_pag_{prefix}_valor").insert(0, fmt_moeda(valor or 0.0))
            getattr(self, f"entry_pag_{prefix}_data").delete(0, "end")
            getattr(self, f"entry_pag_{prefix}_data").insert(0, fmt_data(data or ""))
            getattr(self, f"cb_pag_{prefix}_forma").set(forma)
            getattr(self, f"cb_pag_{prefix}_status").set(status)

        self.btn_excluir.configure(state="normal")
        self._atualizar_lista_itens()

    def _salvar(self):
        try:
            ped = self._coletar_pedido_formulario()
            
            # Regra PD-04: Pré-verificação de estoque
            insumos_negativos = []
            consumo_total = {} # insumo_id -> total_consumo

            for item in ped.itens:
                produto = self.produto_service.get_by_id(item.produto_id)
                if not produto: continue
                for pi in produto.insumos:
                    rendimento = produto.rendimento_receita or 1
                    qtd_usada = (pi.quantidade_usada_receita * item.quantidade) / rendimento
                    consumo_total[pi.insumo_id] = consumo_total.get(pi.insumo_id, 0) + qtd_usada

            for insumo_id, total_consumo in consumo_total.items():
                insumo = self.insumo_service.get_by_id(insumo_id)
                if insumo and (insumo.quantidade_disponivel - total_consumo) < 0:
                    falta = abs(insumo.quantidade_disponivel - total_consumo)
                    insumos_negativos.append(f"{insumo.nome} (Saldo: {insumo.quantidade_disponivel:.2f}, Falta: {falta:.2f})")

            if insumos_negativos:
                lista = "\n- ".join(insumos_negativos)
                if not messagebox.askokcancel("Atenção: Estoque Insuficiente", 
                    f"Os seguintes insumos ficarão com estoque negativo:\n\n- {lista}\n\nDeseja continuar com a gravação do pedido?"):
                    return

            self.pedido_service.salvar(ped)
            messagebox.showinfo("Sucesso", "Pedido salvo com sucesso!")
            self._carregar_clientes(); self._carregar_pedidos(); self._show_lista()
        except (ValueError, Exception) as e:
            messagebox.showerror("Erro", str(e))

    def _excluir(self):
        if not self.current_pedido_id: return
        if messagebox.askyesno("Confirmar exclusão",
                               f"Confirma a exclusão do pedido ID {self.current_pedido_id}?\nEsta ação não pode ser desfeita."):
            try:
                self.pedido_service.excluir(self.current_pedido_id)
                messagebox.showinfo("Sucesso", "Pedido excluído com sucesso!")
                self._carregar_pedidos(); self._show_lista()
            except Exception as e:
                messagebox.showerror("Erro", str(e))

    def _cancelar(self):
        self._show_lista()
