import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.cliente import Cliente
from app.services.cliente_service import ClienteService

class ClientesView(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        
        self.service = ClienteService()
        self.current_id = None
        
        # === Layout Principal ===
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # === Header ===
        header_frame = ctk.CTkFrame(self)
        header_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        self.title_label = ctk.CTkLabel(header_frame, text="Gerenciamento de Clientes", font=("Roboto", 20, "bold"))
        self.title_label.pack(side="left", padx=10, pady=10)
        
        self.btn_novo = ctk.CTkButton(header_frame, text="Novo Cliente", command=self._novo_cliente)
        self.btn_novo.pack(side="right", padx=10, pady=10)

        self.btn_exportar = ctk.CTkButton(header_frame, text="Exportar Excel", command=self._on_exportar_excel)
        self.btn_exportar.pack(side="right", padx=(0, 10), pady=10)
        
        # === Body (Container) ===
        self.body_container = ctk.CTkFrame(self, fg_color="transparent")
        self.body_container.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")
        
        self.body_container.grid_columnconfigure(0, weight=1)
        self.body_container.grid_rowconfigure(0, weight=1)
        
        self.view_lista = ctk.CTkFrame(self.body_container, fg_color="transparent")
        self.view_form = ctk.CTkFrame(self.body_container, fg_color="transparent")
        
        self._setup_lista_view()
        self._setup_form_view()
        
        self._show_lista()
        self._carregar_clientes()

    def _show_lista(self):
        self.view_form.grid_forget()
        self.view_lista.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="normal")
        self.title_label.configure(text="Gerenciamento de Clientes")

    def _show_form(self, editando=False):
        self.view_lista.grid_forget()
        self.view_form.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="disabled")
        self.title_label.configure(text="Editar Cliente" if editando else "Novo Cliente")

    def _formatar_telefone(self, valor: str) -> str:
        digitos = self._apenas_digitos(valor)[:11]
        if len(digitos) <= 2:
            return digitos
        if len(digitos) <= 6:
            return f"({digitos[:2]}) {digitos[2:]}"
        if len(digitos) <= 10:
            return f"({digitos[:2]}) {digitos[2:6]}-{digitos[6:]}"
        return f"({digitos[:2]}) {digitos[2:7]}-{digitos[7:]}"

    def _formatar_cpf_cnpj(self, valor: str) -> str:
        digitos = self._apenas_digitos(valor)[:14]
        if len(digitos) <= 11:
            if len(digitos) <= 3:
                return digitos
            if len(digitos) <= 6:
                return f"{digitos[:3]}.{digitos[3:]}"
            if len(digitos) <= 9:
                return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:]}"
            return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"

        if len(digitos) <= 2:
            return digitos
        if len(digitos) <= 5:
            return f"{digitos[:2]}.{digitos[2:]}"
        if len(digitos) <= 8:
            return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:]}"
        if len(digitos) <= 12:
            return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:]}"
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"

    def _aplicar_mascara_telefone(self, event):
        entry = event.widget
        formatado = self._formatar_telefone(entry.get())
        entry.delete(0, "end")
        entry.insert(0, formatado)

    def _aplicar_mascara_cpf_cnpj(self, event):
        entry = event.widget
        formatado = self._formatar_cpf_cnpj(entry.get())
        entry.delete(0, "end")
        entry.insert(0, formatado)

    def _apenas_digitos(self, valor: str) -> str:
        return "".join(ch for ch in valor if ch.isdigit())

    def _validar_email(self, email: str) -> bool:
        if not email:
            return True
        return re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email) is not None

    def _setup_lista_view(self):
        self.view_lista.grid_columnconfigure(0, weight=1)
        self.view_lista.grid_rowconfigure(1, weight=1)
        
        # Filtros
        filter_frame = ctk.CTkFrame(self.view_lista)
        filter_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        self.filtro_nome = ctk.CTkEntry(filter_frame, placeholder_text="Buscar por Nome...", width=300)
        self.filtro_nome.pack(side="left", padx=10)
        
        self.btn_buscar = ctk.CTkButton(filter_frame, text="Buscar", command=self._carregar_clientes, width=100)
        self.btn_buscar.pack(side="left", padx=10)
        
        # Treeview
        tree_frame = ctk.CTkFrame(self.view_lista)
        tree_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2b2b2b", foreground="white", rowheight=25, fieldbackground="#2b2b2b", borderwidth=0)
        style.map('Treeview', background=[('selected', '#1f538d')])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat")
        
        columns = ("id", "nome", "telefone", "email")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        self.tree.heading("id", text="ID")
        self.tree.heading("nome", text="Nome")
        self.tree.heading("telefone", text="Telefone")
        self.tree.heading("email", text="E-mail")
        
        self.tree.column("id", width=50, anchor="center")
        self.tree.column("nome", width=250)
        self.tree.column("telefone", width=150)
        self.tree.column("email", width=200)
        
        self.tree.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.bind("<Double-1>", self._editar_cliente)

    def _setup_form_view(self):
        form_frame = ctk.CTkFrame(self.view_form)
        form_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(form_frame, text="Nome*:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.entry_nome = ctk.CTkEntry(form_frame, width=300, placeholder_text="Texto")
        self.entry_nome.grid(row=0, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(form_frame, text="Telefone:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        self.entry_telefone = ctk.CTkEntry(form_frame, width=200, placeholder_text="Texto")
        self.entry_telefone.grid(row=1, column=1, padx=10, pady=10, sticky="w")
        self.entry_telefone.bind("<KeyRelease>", self._aplicar_mascara_telefone)
        
        ctk.CTkLabel(form_frame, text="E-mail:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        self.entry_email = ctk.CTkEntry(form_frame, width=300, placeholder_text="Texto")
        self.entry_email.grid(row=2, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(form_frame, text="Endereço:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        self.entry_endereco = ctk.CTkEntry(form_frame, width=300, placeholder_text="Texto")
        self.entry_endereco.grid(row=3, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(form_frame, text="CPF/CNPJ:").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        self.entry_cpf_cnpj = ctk.CTkEntry(form_frame, width=200, placeholder_text="Texto")
        self.entry_cpf_cnpj.grid(row=4, column=1, padx=10, pady=10, sticky="w")
        self.entry_cpf_cnpj.bind("<KeyRelease>", self._aplicar_mascara_cpf_cnpj)
        
        ctk.CTkLabel(form_frame, text="Redes Sociais:").grid(row=5, column=0, padx=10, pady=10, sticky="e")
        self.entry_redes_sociais = ctk.CTkEntry(form_frame, width=300, placeholder_text="Texto")
        self.entry_redes_sociais.grid(row=5, column=1, padx=10, pady=10, sticky="w")
        
        ctk.CTkLabel(form_frame, text="Observações:").grid(row=6, column=0, padx=10, pady=10, sticky="ne")
        self.text_obs = ctk.CTkTextbox(form_frame, width=300, height=80)
        self.text_obs.grid(row=6, column=1, padx=10, pady=10, sticky="w")
        
        # Ações
        acoes_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        acoes_frame.grid(row=7, column=0, columnspan=2, pady=20, sticky="e")
        
        self.btn_salvar = ctk.CTkButton(acoes_frame, text="Salvar", command=self._salvar)
        self.btn_salvar.pack(side="right", padx=5)
        
        self.btn_excluir = ctk.CTkButton(acoes_frame, text="Excluir", fg_color="#c93434", hover_color="#942626", command=self._excluir)
        self.btn_excluir.pack(side="right", padx=5)
        self.btn_excluir.configure(state="disabled")
        
        self.btn_cancelar = ctk.CTkButton(acoes_frame, text="Cancelar", fg_color="gray", hover_color="#555555", command=self._cancelar)
        self.btn_cancelar.pack(side="right", padx=5)

    def _carregar_clientes(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        termo = self.filtro_nome.get()
        clientes = self.service.listar(termo)
        for c in clientes:
            self.tree.insert("", "end", values=(c.id, c.nome, c.telefone, c.email))

    def _on_exportar_excel(self):
        try:
            dados = [self.tree.item(item, "values") for item in self.tree.get_children()]
            if not dados:
                messagebox.showinfo("Exportar Excel", "Não há clientes para exportar com o filtro atual.")
                return

            arquivo = filedialog.asksaveasfilename(
                title="Salvar exportação de clientes",
                defaultextension=".xlsx",
                filetypes=[("Planilha Excel", "*.xlsx")],
                initialfile=f"clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
            if not arquivo:
                return

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Clientes"

            cabecalho = ["ID", "Nome", "Telefone", "E-mail"]
            sheet.append(cabecalho)

            for celula in sheet[1]:
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill(fill_type="solid", fgColor="A66850")
                celula.alignment = Alignment(horizontal="center", vertical="center")

            for linha in dados:
                sheet.append(list(linha))

            larguras = {}
            for linha in sheet.iter_rows():
                for celula in linha:
                    valor = "" if celula.value is None else str(celula.value)
                    larguras[celula.column_letter] = max(larguras.get(celula.column_letter, 0), len(valor))

            for coluna, largura in larguras.items():
                sheet.column_dimensions[coluna].width = min(largura + 2, 45)

            sheet.freeze_panes = "A2"
            workbook.save(arquivo)
            messagebox.showinfo("Exportar Excel", f"Exportação concluída com sucesso.\nArquivo salvo em:\n{arquivo}")
        except Exception as exc:
            messagebox.showerror("Exportar Excel", f"Não foi possível exportar os clientes.\n\n{exc}")

    def _limpar_form(self):
        self.current_id = None
        self.entry_nome.delete(0, 'end')
        self.entry_telefone.delete(0, 'end')
        self.entry_email.delete(0, 'end')
        self.entry_endereco.delete(0, 'end')
        self.entry_cpf_cnpj.delete(0, 'end')
        self.entry_redes_sociais.delete(0, 'end')
        self.text_obs.delete("0.0", "end")
        self.btn_excluir.configure(state="disabled")

    def _novo_cliente(self):
        self._limpar_form()
        self._show_form(editando=False)

    def _editar_cliente(self, event):
        selected = self.tree.selection()
        if not selected:
            return
            
        item = self.tree.item(selected[0])
        self.current_id = item['values'][0]
        
        cliente = self.service.get_by_id(self.current_id)
        if cliente:
            self._limpar_form()
            self.current_id = cliente.id
            self.entry_nome.insert(0, cliente.nome)
            self.entry_telefone.insert(0, self._formatar_telefone(cliente.telefone or ""))
            self.entry_email.insert(0, cliente.email or "")
            self.entry_endereco.insert(0, cliente.endereco or "")
            self.entry_cpf_cnpj.insert(0, self._formatar_cpf_cnpj(cliente.cpf_cnpj or ""))
            self.entry_redes_sociais.insert(0, cliente.redes_sociais or "")
            self.text_obs.insert("0.0", cliente.observacoes or "")
            
            self.btn_excluir.configure(state="normal")
            self._show_form(editando=True)

    def _salvar(self):
        nome = self.entry_nome.get().strip()
        if not nome:
            messagebox.showwarning("Aviso", "O campo Nome é obrigatório.")
            return

        if len(nome) < 3:
            messagebox.showwarning("Aviso", "O campo Nome deve ter ao menos 3 caracteres.")
            return

        telefone = self.entry_telefone.get().strip()
        telefone_digitos = self._apenas_digitos(telefone)
        if telefone and len(telefone_digitos) < 8:
            messagebox.showwarning("Aviso", "Telefone inválido. Informe ao menos 8 dígitos.")
            return

        email = self.entry_email.get().strip()
        if not self._validar_email(email):
            messagebox.showwarning("Aviso", "E-mail inválido.")
            return

        cpf_cnpj = self.entry_cpf_cnpj.get().strip()
        cpf_cnpj_digitos = self._apenas_digitos(cpf_cnpj)
        if cpf_cnpj and len(cpf_cnpj_digitos) not in (11, 14):
            messagebox.showwarning("Aviso", "CPF/CNPJ inválido. Use 11 (CPF) ou 14 (CNPJ) dígitos.")
            return
            
        cli = Cliente(
            id=self.current_id,
            nome=nome,
            telefone=telefone,
            email=email,
            endereco=self.entry_endereco.get().strip(),
            cpf_cnpj=cpf_cnpj,
            redes_sociais=self.entry_redes_sociais.get().strip(),
            observacoes=self.text_obs.get("0.0", "end").strip()
        )
        
        try:
            self.service.salvar(cli)
            messagebox.showinfo("Sucesso", "Cliente salvo com sucesso!")
            self._carregar_clientes()
            self._show_lista()
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def _excluir(self):
        if not self.current_id:
            return

        nome = self.entry_nome.get().strip() or f"ID {self.current_id}"
        mensagem = (
            f"Confirma a exclusao do cliente '{nome}'?\n\n"
            "Esta acao nao pode ser desfeita."
        )
        if messagebox.askyesno("Confirmar exclusao", mensagem):
            try:
                self.service.excluir(self.current_id)
                messagebox.showinfo("Sucesso", "Cliente excluído com sucesso!")
                self._carregar_clientes()
                self._show_lista()
            except Exception as e:
                messagebox.showerror("Erro", str(e))

    def _cancelar(self):
        self._limpar_form()
        self._show_lista()
