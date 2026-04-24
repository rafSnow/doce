import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.formatters import fmt_moeda, parse_float
from app.models.insumo import Insumo
from app.models.produto import Produto
from app.models.produto_insumo import ProdutoInsumo
from app.services.produto_service import ProdutoService
from app.services.insumo_service import InsumoService
from app.core.enums import UnidadeMedida
from app.ui.theme import (
    ACCENT,
    BG_DEEP,
    CARD_BG,
    CARD_BORDER,
    COLOR_GREEN,
    COLOR_RED,
    FIELD_BG,
    HEADER_BG,
    TEXT_MUTED,
    TEXT_PRIMARY,
    TEXT_SECONDARY,
    _treeview_style,
)


class ProdutosView(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=BG_DEEP)
        self.prod_service   = ProdutoService()
        self.insumo_service = InsumoService()

        self.current_produto_id = None
        self.current_insumos_list: List[ProdutoInsumo] = []
        self.matriz_insumos: Dict[int, Insumo] = {}

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # ── Header ────────────────────────────────────────────────────────
        header_frame = ctk.CTkFrame(self, fg_color=HEADER_BG, corner_radius=0, height=56)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(10, 0))
        header_frame.grid_propagate(False)
        header_frame.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            header_frame,
            text="Gerenciamento de Produtos",
            text_color=TEXT_PRIMARY,
            font=ctk.CTkFont(family="Roboto", size=20, weight="bold"),
        )
        self.title_label.grid(row=0, column=0, padx=20, sticky="w")

        actions = ctk.CTkFrame(header_frame, fg_color="transparent")
        actions.grid(row=0, column=1, padx=(6, 16), sticky="e")

        self.btn_novo = ctk.CTkButton(
            actions,
            text="+ Novo Produto",
            command=self._on_novo,
            fg_color=ACCENT,
            hover_color="#A84F5E",
            text_color="#FFFFFF",
            height=30, corner_radius=20,
            font=ctk.CTkFont(size=12, weight="bold"),
        )
        self.btn_novo.grid(row=0, column=0, padx=4)

        self.btn_exportar = ctk.CTkButton(
            actions,
            text="Exportar Excel",
            command=self._on_exportar_excel,
            fg_color=FIELD_BG,
            hover_color="#F2D5DC",
            border_color=CARD_BORDER, border_width=1,
            text_color=TEXT_SECONDARY,
            height=30, corner_radius=20,
            font=ctk.CTkFont(size=12),
        )
        self.btn_exportar.grid(row=0, column=1, padx=4)

        ctk.CTkFrame(self, fg_color=CARD_BORDER, height=1).grid(row=1, column=0, sticky="ew")

        # ── Body ──────────────────────────────────────────────────────────
        self.body_container = ctk.CTkFrame(self, fg_color="transparent")
        self.body_container.grid(row=2, column=0, padx=14, pady=(10, 12), sticky="nsew")
        self.body_container.grid_columnconfigure(0, weight=1)
        self.body_container.grid_rowconfigure(0, weight=1)

        self.view_lista = ctk.CTkFrame(self.body_container, fg_color="transparent")
        self.view_form  = ctk.CTkFrame(self.body_container, fg_color="transparent")

        self._build_table()
        self._build_form_sidebar()

        self._show_lista()
        self._carregar_dados()

    def refresh(self):
        self._carregar_dados()

    # ── visibilidade ──────────────────────────────────────────────────────
    def _show_lista(self):
        self.view_form.grid_forget()
        self.view_lista.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="normal")
        self.title_label.configure(text="Gerenciamento de Produtos")

    def _show_form(self, editando=False):
        self.view_lista.grid_forget()
        self.view_form.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="disabled")
        self.title_label.configure(text="Editar Produto" if editando else "Novo Produto")
        self.entry_nome.focus_set()

    def _aplicar_mascara_decimal(self, event):
        entry = event.widget
        raw = entry.get().replace(",", ".")
        resultado, tem_ponto = [], False
        for ch in raw:
            if ch.isdigit():
                resultado.append(ch)
            elif ch == "." and not tem_ponto:
                resultado.append(ch); tem_ponto = True
        sanitizado = "".join(resultado)
        if "." in sanitizado:
            inteiro, decimal = sanitizado.split(".", 1)
            sanitizado = f"{inteiro}.{decimal[:2]}"
        entry.delete(0, "end"); entry.insert(0, sanitizado)

    def _aplicar_mascara_inteiro(self, event):
        entry = event.widget
        sanitizado = "".join(ch for ch in entry.get() if ch.isdigit())
        entry.delete(0, "end"); entry.insert(0, sanitizado)

    # ── tabela ────────────────────────────────────────────────────────────
    def _build_table(self):
        self.view_lista.grid_columnconfigure(0, weight=1)
        self.view_lista.grid_rowconfigure(1, weight=1)

        # filtros
        filter_frame = ctk.CTkFrame(
            self.view_lista, fg_color=CARD_BG, corner_radius=12,
            border_width=1, border_color=CARD_BORDER,
        )
        filter_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(filter_frame, text="Filtros", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold"),
                     ).pack(side="left", padx=(12, 6), pady=10)

        self.entry_busca = ctk.CTkEntry(
            filter_frame,
            placeholder_text="Buscar por nome...", width=340,
            fg_color=FIELD_BG, border_color=CARD_BORDER, text_color=TEXT_PRIMARY,
        )
        self.entry_busca.pack(side="left", padx=10)
        self.entry_busca.bind("<KeyRelease>", lambda e: self._carregar_dados())

        # tabela
        table_frame = ctk.CTkFrame(
            self.view_lista, fg_color=CARD_BG, corner_radius=12,
            border_width=1, border_color=CARD_BORDER,
        )
        table_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        _treeview_style("Produtos", rowheight=27)

        columns = ("id", "nome", "rendimento", "custo", "comissao", "preco", "pode_produzir")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings",
                                 style="Produtos.Treeview")

        self.tree.heading("id",         text="ID")
        self.tree.heading("nome",       text="Nome")
        self.tree.heading("rendimento", text="Rendimento")
        self.tree.heading("custo",      text="Custo Unita. (R$)")
        self.tree.heading("comissao",   text="Markup (%)")
        self.tree.heading("preco",      text="Preço Venda (R$)")
        self.tree.heading("pode_produzir", text="Pode Produzir")

        self.tree.column("id",         width=40,  anchor="center")
        self.tree.column("nome",       width=180, anchor="w")
        self.tree.column("rendimento", width=90,  anchor="center")
        self.tree.column("custo",      width=110, anchor="center")
        self.tree.column("comissao",   width=90,  anchor="center")
        self.tree.column("preco",      width=110, anchor="center")
        self.tree.column("pode_produzir", width=100, anchor="center")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(fill="both", expand=True, side="left")
        scrollbar.pack(fill="y", side="right")

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Button-3>", self._show_context_menu)

        self.context_menu = tk.Menu(self, tearoff=0,
                                    bg=CARD_BG, fg=TEXT_PRIMARY,
                                    activebackground="#FAE8EC", activeforeground=ACCENT)
        self.context_menu.add_command(label="Editar",    command=self._on_editar_selecionado)
        self.context_menu.add_command(label="Duplicar",  command=self._on_duplicar_selecionado)
        self.context_menu.add_command(label="Excluir",   command=self._on_excluir_selecionado)

    # ── formulário ────────────────────────────────────────────────────────
    def _build_form_sidebar(self):
        self.form_frame = ctk.CTkFrame(
            self.view_form, fg_color=CARD_BG, corner_radius=12,
            border_width=1, border_color=CARD_BORDER,
        )
        self.form_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Atalhos de teclado
        self.view_form.bind("<Control-s>", lambda e: self._on_salvar())
        self.view_form.bind("<Escape>",    lambda e: self._ocultar_form())

        inner_form = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        inner_form.pack(expand=True, padx=20, pady=20)
        inner_form.grid_columnconfigure(0, weight=1)
        inner_form.grid_columnconfigure(1, weight=1)

        # ── Dados Base ────────────────────────────────────────────────────
        ctk.CTkLabel(inner_form, text="Dados do Produto",
                     text_color=TEXT_MUTED, font=ctk.CTkFont(size=10, weight="bold"),
                     ).grid(row=0, column=0, columnspan=2, pady=10, sticky="w")

        ctk.CTkLabel(inner_form, text="Nome*", text_color=TEXT_SECONDARY
                     ).grid(row=1, column=0, padx=10, pady=(5, 2), sticky="w")
        self.entry_nome = ctk.CTkEntry(
            inner_form, placeholder_text="Texto", width=300,
            fg_color=FIELD_BG, border_color=CARD_BORDER, text_color=TEXT_PRIMARY,
        )
        self.entry_nome.grid(row=2, column=0, columnspan=2, padx=10, pady=(0, 8), sticky="ew")

        ctk.CTkLabel(inner_form, text="Rendimento*",  text_color=TEXT_SECONDARY
                     ).grid(row=3, column=0, padx=10, pady=(5, 2), sticky="w")
        ctk.CTkLabel(inner_form, text="Markup (%)*",  text_color=TEXT_SECONDARY
                     ).grid(row=3, column=1, padx=10, pady=(5, 2), sticky="w")

        self.entry_rendimento = ctk.CTkEntry(
            inner_form, placeholder_text="Numero",
            fg_color=FIELD_BG, border_color=CARD_BORDER, text_color=TEXT_PRIMARY,
        )
        self.entry_rendimento.grid(row=4, column=0, padx=10, pady=(0, 8), sticky="ew")

        self.entry_markup = ctk.CTkEntry(
            inner_form, placeholder_text="Numero",
            fg_color=FIELD_BG, border_color=CARD_BORDER, text_color=TEXT_PRIMARY,
        )
        self.entry_markup.grid(row=4, column=1, padx=10, pady=(0, 8), sticky="ew")

        self.entry_rendimento.bind("<KeyRelease>",
            lambda e: (self._aplicar_mascara_inteiro(e), self._atualizar_calculos_label()))
        self.entry_markup.bind("<KeyRelease>",
            lambda e: (self._aplicar_mascara_decimal(e), self._atualizar_calculos_label()))

        # ── Insumos ───────────────────────────────────────────────────────
        ctk.CTkLabel(inner_form, text="Insumos (Ficha Técnica)",
                     text_color=TEXT_MUTED, font=ctk.CTkFont(size=10, weight="bold"),
                     ).grid(row=5, column=0, columnspan=2, pady=(15, 5), sticky="w")

        frame_add = ctk.CTkFrame(inner_form, fg_color="transparent")
        frame_add.grid(row=6, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        frame_add.grid_columnconfigure(0, weight=1)
        frame_add.grid_columnconfigure(1, weight=1)
        frame_add.grid_columnconfigure(2, weight=0)

        ctk.CTkLabel(frame_add, text="Insumo",     text_color=TEXT_SECONDARY
                     ).grid(row=0, column=0, padx=(0, 6), pady=(0, 2), sticky="w")
        ctk.CTkLabel(frame_add, text="Quantidade", text_color=TEXT_SECONDARY
                     ).grid(row=0, column=1, padx=(0, 6), pady=(0, 2), sticky="w")

        self.combo_insumo = ctk.CTkComboBox(
            frame_add, values=["Carregando..."],
            fg_color=FIELD_BG, border_color=CARD_BORDER,
            button_color=CARD_BORDER, button_hover_color="#E0B8C2",
            text_color=TEXT_PRIMARY,
            dropdown_fg_color=CARD_BG, dropdown_text_color=TEXT_PRIMARY,
            dropdown_hover_color="#FAE8EC",
        )
        self.combo_insumo.grid(row=1, column=0, padx=(0, 6), pady=(0, 5), sticky="ew")

        self.entry_qtd_insumo = ctk.CTkEntry(
            frame_add, placeholder_text="Numero", width=80,
            fg_color=FIELD_BG, border_color=CARD_BORDER, text_color=TEXT_PRIMARY,
        )
        self.entry_qtd_insumo.grid(row=1, column=1, padx=(0, 6), pady=(0, 5), sticky="ew")
        self.entry_qtd_insumo.bind("<KeyRelease>", self._aplicar_mascara_decimal)

        ctk.CTkButton(
            frame_add, text="Adicionar", width=80,
            command=self._adicionar_insumo_na_lista,
            fg_color=ACCENT, hover_color="#A84F5E", text_color="#FFFFFF", corner_radius=10,
        ).grid(row=1, column=2, pady=(0, 5), sticky="e")

        # treeview ficha técnica
        tree_frame = ctk.CTkFrame(
            inner_form, fg_color=FIELD_BG, corner_radius=10,
            border_width=1, border_color=CARD_BORDER,
        )
        tree_frame.grid(row=7, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        inner_form.grid_rowconfigure(7, weight=1)

        cols = ("id", "nome", "qtd", "custo")
        self.tree_insumos = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                         height=5, style="Produtos.Treeview")
        self.tree_insumos.heading("id",    text="ID")
        self.tree_insumos.heading("nome",  text="Insumo")
        self.tree_insumos.heading("qtd",   text="Qtd")
        self.tree_insumos.heading("custo", text="Custo R$")
        self.tree_insumos.column("id",    width=0, stretch=False)
        self.tree_insumos.column("nome",  width=120, anchor="w")
        self.tree_insumos.column("qtd",   width=50,  anchor="center")
        self.tree_insumos.column("custo", width=60,  anchor="center")
        self.tree_insumos.pack(fill="both", expand=True)

        self.tree_insumos.bind("<Delete>",   self._remover_insumo_selecionado)
        self.tree_insumos.bind("<BackSpace>", self._remover_insumo_selecionado)
        self.tree_insumos.bind("<Double-1>",  self._remover_insumo_selecionado)

        # totais
        frame_totais = ctk.CTkFrame(
            inner_form, fg_color=HEADER_BG, corner_radius=10,
            border_width=1, border_color=CARD_BORDER,
        )
        frame_totais.grid(row=8, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        self.lbl_custo_receita = ctk.CTkLabel(
            frame_totais, text="Custo Receita: R$ 0.00", text_color=TEXT_PRIMARY)
        self.lbl_custo_receita.pack(anchor="w", padx=10, pady=2)

        self.lbl_custo_un = ctk.CTkLabel(
            frame_totais, text="Custo Unita.: R$ 0.00", text_color=COLOR_RED)
        self.lbl_custo_un.pack(anchor="w", padx=10, pady=2)

        self.lbl_preco_venda = ctk.CTkLabel(
            frame_totais, text="Preço Venda: R$ 0.00",
            text_color=COLOR_GREEN, font=ctk.CTkFont(weight="bold"))
        self.lbl_preco_venda.pack(anchor="w", padx=10, pady=2)

        # botões
        frame_btns = ctk.CTkFrame(inner_form, fg_color="transparent")
        frame_btns.grid(row=9, column=0, columnspan=2, padx=10, pady=10, sticky="e")

        self.btn_salvar = ctk.CTkButton(
            frame_btns, text="Salvar", command=self._on_salvar,
            fg_color=ACCENT, hover_color="#A84F5E", text_color="#FFFFFF", corner_radius=10,
        )
        self.btn_salvar.pack(side="right", padx=5)

        self.btn_excluir = ctk.CTkButton(
            frame_btns, text="Excluir",
            fg_color="#FAE8EC", hover_color="#F2D5DC",
            border_color=CARD_BORDER, border_width=1,
            text_color=ACCENT, corner_radius=10,
            command=self._excluir_form,
        )
        self.btn_excluir.pack(side="right", padx=5)
        self.btn_excluir.configure(state="disabled")

        self.btn_cancelar = ctk.CTkButton(
            frame_btns, text="Cancelar",
            fg_color=FIELD_BG, hover_color="#F2D5DC",
            border_color=CARD_BORDER, border_width=1,
            text_color=TEXT_SECONDARY, corner_radius=10,
            command=self._ocultar_form,
        )
        self.btn_cancelar.pack(side="right", padx=5)

        ctk.CTkLabel(inner_form, text="* Campos obrigatórios", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10)).grid(row=10, column=0, columnspan=2, padx=10, pady=(0, 5), sticky="w")

    # ── dados ─────────────────────────────────────────────────────────────
    def _carregar_dados(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for p in self.prod_service.listar(nome=self.entry_busca.get()):
            capacidade = self.insumo_service.calcular_capacidade_producao(p.id)
            self.tree.insert("", "end", values=(
                p.id, p.nome, p.rendimento_receita,
                fmt_moeda(p.custo_unitario),
                f"{p.comissao_perc}%",
                fmt_moeda(p.preco_venda_unitario),
                capacidade,
            ))

    def _carregar_combo_insumos(self):
        insumos = self.insumo_service.listar()
        self.matriz_insumos = {i.id: i for i in insumos}
        combo_vals = [f"{i.id} - {i.nome} ({i.unidade_medida})" for i in insumos]
        self.combo_insumo.configure(values=combo_vals)
        self.combo_insumo.set(combo_vals[0] if combo_vals else "Nenhum insumo.")

    def _on_novo(self):
        self.current_produto_id = None
        self.current_insumos_list.clear()
        self.entry_nome.delete(0, "end")
        self.entry_rendimento.delete(0, "end"); self.entry_rendimento.insert(0, "1")
        
        from app.services.configuracao_service import ConfiguracaoService
        markup_padrao = ConfiguracaoService().get_markup_padrao()
        self.entry_markup.delete(0, "end");    self.entry_markup.insert(0, str(markup_padrao))
        
        self.btn_excluir.configure(state="disabled")
        self._carregar_combo_insumos()
        self._render_tree_insumos()
        self._show_form()

    def _ocultar_form(self):
        self._show_lista()

    def _adicionar_insumo_na_lista(self):
        selecao = self.combo_insumo.get()
        qtd_txt = self.entry_qtd_insumo.get().strip().replace(",", ".")
        if "-" not in selecao:
            messagebox.showerror("Erro", "Selecione um insumo válido."); return
        if not qtd_txt:
            messagebox.showerror("Erro", "Informe a quantidade usada."); return
        insumo_id = int(selecao.split(" - ")[0])
        try:
            qtd = float(qtd_txt)
        except ValueError:
            messagebox.showerror("Erro", "Quantidade inválida."); return
        if qtd <= 0:
            messagebox.showerror("Erro", "A quantidade deve ser maior que zero."); return
        existente = next((pi for pi in self.current_insumos_list if pi.insumo_id == insumo_id), None)
        if existente:
            existente.quantidade_usada_receita += qtd
        else:
            insumo_ref = self.matriz_insumos.get(insumo_id)
            if not insumo_ref:
                messagebox.showerror("Erro", "Insumo não encontrado."); return
            self.current_insumos_list.append(ProdutoInsumo(
                insumo_id=insumo_id,
                quantidade_usada_receita=qtd,
                insumo_nome=insumo_ref.nome,
                insumo_unidade=insumo_ref.unidade_medida,
            ))
        self.entry_qtd_insumo.delete(0, "end")
        self._render_tree_insumos()

    def _remover_insumo_selecionado(self, event=None):
        selected = self.tree_insumos.selection()
        if not selected: return
        insumo_id = int(self.tree_insumos.item(selected[0])["values"][0])
        self.current_insumos_list = [pi for pi in self.current_insumos_list if pi.insumo_id != insumo_id]
        self._render_tree_insumos()

    def _render_tree_insumos(self):
        for row in self.tree_insumos.get_children():
            self.tree_insumos.delete(row)
        for pi in self.current_insumos_list:
            insumo_ref = self.matriz_insumos.get(pi.insumo_id)
            custo_prop = (pi.quantidade_usada_receita * insumo_ref.custo_por_unidade
                          if insumo_ref else 0.0)
            self.tree_insumos.insert("", "end", values=(
                pi.insumo_id,
                pi.insumo_nome,
                f"{pi.quantidade_usada_receita:g} {pi.insumo_unidade}",
                fmt_moeda(custo_prop),
            ))
        self._atualizar_calculos_label()

    def _atualizar_calculos_label(self, event=None):
        try:
            rendimento = int(self.entry_rendimento.get() or 1)
            markup = float(self.entry_markup.get().replace(",", ".") or 0)
        except ValueError:
            rendimento, markup = 1, 0.0
        custo_total = sum(
            pi.quantidade_usada_receita * self.matriz_insumos[pi.insumo_id].custo_por_unidade
            for pi in self.current_insumos_list
            if pi.insumo_id in self.matriz_insumos
        )
        custo_un    = (custo_total / rendimento) if rendimento > 0 else 0
        preco_venda = custo_un * (1 + (markup / 100))
        
        # Regra P-07: Alerta visual de margem negativa em tempo real
        status_venda = ""
        cor_venda    = COLOR_GREEN
        if preco_venda < custo_un:
            status_venda = " (ABAIXO DO CUSTO!)"
            cor_venda    = COLOR_RED

        self.lbl_custo_receita.configure(text=f"Custo Receita: R$ {custo_total:.2f}")
        self.lbl_custo_un.configure(text=f"Custo Unita.: R$ {custo_un:.2f}")
        self.lbl_preco_venda.configure(
            text=f"Preço Venda: R$ {preco_venda:.2f}{status_venda}",
            text_color=cor_venda
        )

    def _on_salvar(self):
        nome = self.entry_nome.get().strip()
        try:
            if not nome:           raise ValueError("Nome é obrigatório.")
            if len(nome) < 3:      raise ValueError("Nome deve ter ao menos 3 caracteres.")
            rend_txt = self.entry_rendimento.get().strip()
            if not rend_txt:       raise ValueError("Rendimento é obrigatório.")
            rend = int(rend_txt)
            if rend <= 0:          raise ValueError("Rendimento deve ser maior que zero.")
            mkup = parse_float(self.entry_markup.get(), "Markup", minimo=0.0)
            if not self.current_insumos_list:
                raise ValueError("Adicione ao menos um insumo na ficha técnica.")
            for pi in self.current_insumos_list:
                if pi.quantidade_usada_receita <= 0:
                    raise ValueError("Todas as quantidades de insumos devem ser maiores que zero.")
        except ValueError as e:
            messagebox.showerror("Erro", str(e)); return

        # Regra P-07: Alerta de margem negativa antes de persistir
        custo_total = sum(
            pi.quantidade_usada_receita * self.matriz_insumos[pi.insumo_id].custo_por_unidade
            for pi in self.current_insumos_list
            if pi.insumo_id in self.matriz_insumos
        )
        custo_un = custo_total / rend if rend > 0 else 0
        preco_venda = custo_un * (1 + (mkup / 100))

        if preco_venda < custo_un:
            if not messagebox.askokcancel("Aviso de Margem",
                f"Atenção: O preço de venda (R$ {preco_venda:.2f}) está abaixo do custo unitário (R$ {custo_un:.2f}).\n"
                "Deseja salvar mesmo assim?"):
                return

        self.prod_service.salvar(Produto(
            id=self.current_produto_id, nome=nome,
            rendimento_receita=rend, comissao_perc=mkup,
            insumos=self.current_insumos_list,
        ))
        self._ocultar_form(); self._carregar_dados()

    # ── context menu / double click ───────────────────────────────────────
    def _show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.tk_popup(event.x_root, event.y_root)

    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if item: self._on_editar_selecionado()

    def _on_editar_selecionado(self):
        selected = self.tree.selection()
        if not selected: return
        prod = self.prod_service.get_by_id(self.tree.item(selected[0])["values"][0])
        if not prod: return
        self.current_produto_id = prod.id
        self.entry_nome.delete(0, "end"); self.entry_nome.insert(0, prod.nome)
        self.entry_rendimento.delete(0, "end"); self.entry_rendimento.insert(0, str(prod.rendimento_receita))
        self.entry_markup.delete(0, "end"); self.entry_markup.insert(0, str(prod.comissao_perc))
        self.btn_excluir.configure(state="normal")
        self._carregar_combo_insumos()
        self.current_insumos_list = prod.insumos
        self._render_tree_insumos()
        self._show_form(editando=True)

    def _excluir_form(self):
        if not self.current_produto_id: return
        nome = self.entry_nome.get()
        if messagebox.askyesno("Confirmar exclusão",
                               f"Confirma a exclusão do produto '{nome}'?\nEsta ação não pode ser desfeita."):
            self.prod_service.excluir(self.current_produto_id)
            self._carregar_dados(); self._ocultar_form()

    def _on_duplicar_selecionado(self):
        selected = self.tree.selection()
        if not selected: return
        self.prod_service.duplicar(self.tree.item(selected[0])["values"][0])
        self._carregar_dados()

    def _on_excluir_selecionado(self):
        selected = self.tree.selection()
        if not selected: return
        item_id = self.tree.item(selected[0])["values"][0]
        nome    = self.tree.item(selected[0])["values"][1]
        if messagebox.askyesno("Confirmar exclusão",
                               f"Confirma a exclusão do produto '{nome}'?\nEsta ação não pode ser desfeita."):
            self.prod_service.excluir(item_id)
            self._carregar_dados()
            if self.current_produto_id == item_id:
                self._ocultar_form()

    def _on_exportar_excel(self):
        produtos = self.prod_service.listar(nome=self.entry_busca.get().strip())
        if not produtos:
            messagebox.showinfo("Exportar Excel", "Não há produtos para exportar."); return
        caminho = filedialog.asksaveasfilename(
            title="Salvar exportação de produtos", defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            initialfile=f"produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not caminho: return
        try:
            wb = Workbook()
            ws_res = wb.active; ws_res.title = "Produtos"
            cab_res = ["ID","Nome","Rendimento","Custo Unitário (R$)","Markup (%)","Preço de Venda (R$)"]
            ws_res.append(cab_res)
            for c in ws_res[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C96B7A")
                c.alignment = Alignment(horizontal="center", vertical="center")

            ws_fic = wb.create_sheet("Ficha Tecnica")
            cab_fic = ["Produto ID","Produto","Insumo ID","Insumo",
                       "Quantidade Usada","Unidade","Custo Proporcional (R$)"]
            ws_fic.append(cab_fic)
            for c in ws_fic[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="7A4A55")
                c.alignment = Alignment(horizontal="center", vertical="center")

            for p in produtos:
                ws_res.append([p.id, p.nome, p.rendimento_receita,
                               p.custo_unitario, p.comissao_perc, p.preco_venda_unitario])
                pd = self.prod_service.get_by_id(p.id)
                if pd:
                    for pi in pd.insumos:
                        ws_fic.append([pd.id, pd.nome, pi.insumo_id, pi.insumo_nome or "",
                                       pi.quantidade_usada_receita, pi.insumo_unidade or "",
                                       pi.custo_proporcional])

            for planilha in (ws_res, ws_fic):
                for col in planilha.columns:
                    planilha.column_dimensions[col[0].column_letter].width = min(
                        max(len(str(c.value or "")) for c in col) + 2, 40)
                planilha.freeze_panes = "A2"

            wb.save(caminho)
            messagebox.showinfo("Exportar Excel", f"Exportação concluída.\nArquivo salvo em:\n{caminho}")
        except Exception as exc:
            messagebox.showerror("Exportar Excel", f"Não foi possível exportar.\n\n{exc}")
