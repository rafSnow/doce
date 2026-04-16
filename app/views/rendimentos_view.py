import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.rendimento import Rendimento
from app.services.cliente_service import ClienteService
from app.services.rendimento_service import RendimentoService


class RendimentosView(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Financeiro - Rendimentos")
        self.geometry("980x640")
        self.minsize(860, 560)

        self.service = RendimentoService()
        self.cliente_service = ClienteService()
        self.current_rendimento_id = None
        self.cliente_map: dict[int, str] = {}

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header_frame = ctk.CTkFrame(self)
        header_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        self.title_label = ctk.CTkLabel(header_frame, text="Financeiro - Rendimentos", font=("Roboto", 20, "bold"))
        self.title_label.pack(side="left", padx=10, pady=10)

        self.btn_exportar = ctk.CTkButton(header_frame, text="Exportar Excel", command=self._on_exportar_excel)
        self.btn_exportar.pack(side="right", padx=(0, 10), pady=10)

        self.btn_novo = ctk.CTkButton(header_frame, text="Novo Rendimento", command=self._on_novo)
        self.btn_novo.pack(side="right", padx=10, pady=10)

        self.body_container = ctk.CTkFrame(self, fg_color="transparent")
        self.body_container.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")
        self.body_container.grid_columnconfigure(0, weight=1)
        self.body_container.grid_rowconfigure(0, weight=1)

        self.view_lista = ctk.CTkFrame(self.body_container, fg_color="transparent")
        self.view_form = ctk.CTkFrame(self.body_container, fg_color="transparent")

        self._build_table()
        self._build_form()

        self._show_lista()
        self._carregar_clientes()
        self._carregar_dados_seguro()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.grab_set()

    def _on_close(self):
        self.grab_release()
        self.destroy()

    def _show_lista(self):
        self.view_form.grid_forget()
        self.view_lista.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="normal")
        self.title_label.configure(text="Financeiro - Rendimentos")

    def _show_form(self, editando: bool = False):
        self.view_lista.grid_forget()
        self.view_form.grid(row=0, column=0, sticky="nsew")
        self.btn_novo.configure(state="disabled")
        self.title_label.configure(text="Editar Rendimento" if editando else "Novo Rendimento")

    def _build_table(self):
        self.view_lista.grid_columnconfigure(0, weight=1)
        self.view_lista.grid_rowconfigure(1, weight=1)

        filtro_frame = ctk.CTkFrame(self.view_lista)
        filtro_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))

        ctk.CTkLabel(filtro_frame, text="Período:").pack(side="left", padx=(10, 4), pady=10)

        self.filtro_data_inicio = ctk.CTkEntry(filtro_frame, width=110, placeholder_text="DD/MM/AAAA")
        self.filtro_data_inicio.pack(side="left", padx=4, pady=10)
        self.filtro_data_inicio.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(filtro_frame, text="até").pack(side="left", padx=4, pady=10)

        self.filtro_data_fim = ctk.CTkEntry(filtro_frame, width=110, placeholder_text="DD/MM/AAAA")
        self.filtro_data_fim.pack(side="left", padx=4, pady=10)
        self.filtro_data_fim.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(filtro_frame, text="Status:").pack(side="left", padx=(14, 4), pady=10)
        self.filtro_status = ctk.CTkOptionMenu(filtro_frame, values=["Todos", "Pendente", "Recebido"])
        self.filtro_status.pack(side="left", padx=4, pady=10)
        self.filtro_status.set("Todos")

        self.btn_filtrar = ctk.CTkButton(filtro_frame, text="Buscar", width=90, command=self._carregar_dados_seguro)
        self.btn_filtrar.pack(side="left", padx=(14, 4), pady=10)

        self.btn_limpar_filtros = ctk.CTkButton(
            filtro_frame,
            text="Limpar",
            width=90,
            fg_color="gray",
            hover_color="#555555",
            command=self._limpar_filtros,
        )
        self.btn_limpar_filtros.pack(side="left", padx=4, pady=10)

        table_frame = ctk.CTkFrame(self.view_lista)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2b2b2b", foreground="white", rowheight=25, fieldbackground="#2b2b2b", borderwidth=0)
        style.map("Treeview", background=[("selected", "#1f538d")])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat")

        columns = ("id", "cliente", "pag_inicial", "status_inicial", "pag_final", "status_final", "responsavel")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")

        self.tree.heading("id", text="ID")
        self.tree.heading("cliente", text="Cliente")
        self.tree.heading("pag_inicial", text="Pag. Inicial")
        self.tree.heading("status_inicial", text="Status Inicial")
        self.tree.heading("pag_final", text="Pag. Final")
        self.tree.heading("status_final", text="Status Final")
        self.tree.heading("responsavel", text="Responsável")

        self.tree.column("id", width=50, anchor="center")
        self.tree.column("cliente", width=220, anchor="w")
        self.tree.column("pag_inicial", width=120, anchor="e")
        self.tree.column("status_inicial", width=110, anchor="center")
        self.tree.column("pag_final", width=120, anchor="e")
        self.tree.column("status_final", width=110, anchor="center")
        self.tree.column("responsavel", width=150, anchor="w")

        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.bind("<Double-1>", self._on_editar_selecionado)

    def _build_form(self):
        form_frame = ctk.CTkFrame(self.view_form)
        form_frame.pack(fill="both", expand=True, padx=20, pady=20)
        form_frame.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(form_frame, text="Cliente*").grid(row=0, column=0, padx=10, pady=8, sticky="e")
        self.cb_cliente = ctk.CTkComboBox(form_frame, values=["Carregando clientes..."])
        self.cb_cliente.grid(row=0, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(form_frame, text="Responsável").grid(row=0, column=2, padx=10, pady=8, sticky="e")
        self.entry_responsavel = ctk.CTkEntry(form_frame, placeholder_text="Texto")
        self.entry_responsavel.grid(row=0, column=3, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(form_frame, text="Valor Inicial (R$)").grid(row=1, column=0, padx=10, pady=8, sticky="e")
        self.entry_pag_ini_valor = ctk.CTkEntry(form_frame, placeholder_text="0,00")
        self.entry_pag_ini_valor.grid(row=1, column=1, padx=10, pady=8, sticky="ew")
        self.entry_pag_ini_valor.bind("<KeyRelease>", self._aplicar_mascara_moeda)

        ctk.CTkLabel(form_frame, text="Data Inicial").grid(row=1, column=2, padx=10, pady=8, sticky="e")
        self.entry_pag_ini_data = ctk.CTkEntry(form_frame, placeholder_text="DD/MM/AAAA")
        self.entry_pag_ini_data.grid(row=1, column=3, padx=10, pady=8, sticky="ew")
        self.entry_pag_ini_data.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(form_frame, text="Forma Inicial").grid(row=2, column=0, padx=10, pady=8, sticky="e")
        self.cb_pag_ini_forma = ctk.CTkComboBox(form_frame, values=["Pix", "Dinheiro", "Cartão Crédito", "Cartão Débito", "Transferência"])
        self.cb_pag_ini_forma.grid(row=2, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(form_frame, text="Status Inicial").grid(row=2, column=2, padx=10, pady=8, sticky="e")
        self.cb_pag_ini_status = ctk.CTkComboBox(form_frame, values=["Pendente", "Recebido"])
        self.cb_pag_ini_status.grid(row=2, column=3, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(form_frame, text="Valor Final (R$)").grid(row=3, column=0, padx=10, pady=8, sticky="e")
        self.entry_pag_fin_valor = ctk.CTkEntry(form_frame, placeholder_text="0,00")
        self.entry_pag_fin_valor.grid(row=3, column=1, padx=10, pady=8, sticky="ew")
        self.entry_pag_fin_valor.bind("<KeyRelease>", self._aplicar_mascara_moeda)

        ctk.CTkLabel(form_frame, text="Data Final").grid(row=3, column=2, padx=10, pady=8, sticky="e")
        self.entry_pag_fin_data = ctk.CTkEntry(form_frame, placeholder_text="DD/MM/AAAA")
        self.entry_pag_fin_data.grid(row=3, column=3, padx=10, pady=8, sticky="ew")
        self.entry_pag_fin_data.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(form_frame, text="Forma Final").grid(row=4, column=0, padx=10, pady=8, sticky="e")
        self.cb_pag_fin_forma = ctk.CTkComboBox(form_frame, values=["Pix", "Dinheiro", "Cartão Crédito", "Cartão Débito", "Transferência"])
        self.cb_pag_fin_forma.grid(row=4, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(form_frame, text="Status Final").grid(row=4, column=2, padx=10, pady=8, sticky="e")
        self.cb_pag_fin_status = ctk.CTkComboBox(form_frame, values=["Pendente", "Recebido"])
        self.cb_pag_fin_status.grid(row=4, column=3, padx=10, pady=8, sticky="ew")

        frame_acoes = ctk.CTkFrame(form_frame, fg_color="transparent")
        frame_acoes.grid(row=5, column=0, columnspan=4, padx=10, pady=16, sticky="e")

        self.btn_salvar = ctk.CTkButton(frame_acoes, text="Salvar", command=self._on_salvar)
        self.btn_salvar.pack(side="right", padx=5)

        self.btn_excluir = ctk.CTkButton(frame_acoes, text="Excluir", fg_color="#c93434", hover_color="#942626", command=self._on_excluir)
        self.btn_excluir.pack(side="right", padx=5)

        self.btn_cancelar = ctk.CTkButton(frame_acoes, text="Cancelar", fg_color="gray", hover_color="#555555", command=self._show_lista)
        self.btn_cancelar.pack(side="right", padx=5)

        self._limpar_form()

    def _carregar_clientes(self):
        self.cliente_map.clear()
        clientes = self.cliente_service.listar()
        labels = []
        for cliente in clientes:
            if cliente.id is None:
                continue
            label = f"{cliente.id} - {cliente.nome}"
            self.cliente_map[int(cliente.id)] = label
            labels.append(label)

        if not labels:
            labels = ["Nenhum cliente cadastrado"]

        self.cb_cliente.configure(values=labels)
        self.cb_cliente.set(labels[0])

    def _carregar_dados(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        data_inicio = self._validar_data(self.filtro_data_inicio.get().strip(), "Data inicial")
        data_fim = self._validar_data(self.filtro_data_fim.get().strip(), "Data final")

        if data_inicio and data_fim:
            inicio_dt = datetime.strptime(data_inicio, "%d/%m/%Y")
            fim_dt = datetime.strptime(data_fim, "%d/%m/%Y")
            if inicio_dt > fim_dt:
                raise ValueError("A data inicial não pode ser maior que a data final.")

        status = self.filtro_status.get().strip() if hasattr(self, "filtro_status") else "Todos"

        rendimentos = self.service.listar(
            data_inicio=data_inicio,
            data_fim=data_fim,
            status=status,
        )
        for rendimento in rendimentos:
            cliente_nome = self.cliente_map.get(rendimento.cliente_id, "Cliente não encontrado")
            self.tree.insert(
                "",
                "end",
                values=(
                    rendimento.id,
                    cliente_nome,
                    f"R$ {self._formatar_moeda_ui(rendimento.pag_inicial_valor or 0.0)}",
                    rendimento.pag_inicial_status,
                    f"R$ {self._formatar_moeda_ui(rendimento.pag_final_valor or 0.0)}",
                    rendimento.pag_final_status,
                    rendimento.responsavel or "",
                ),
            )

    def _carregar_dados_seguro(self):
        try:
            self._carregar_dados()
        except ValueError as exc:
            messagebox.showerror("Erro", str(exc))

    def _limpar_filtros(self):
        self.filtro_data_inicio.delete(0, "end")
        self.filtro_data_fim.delete(0, "end")
        self.filtro_status.set("Todos")
        self._carregar_dados_seguro()

    def _on_exportar_excel(self):
        try:
            data_inicio = self._validar_data(self.filtro_data_inicio.get().strip(), "Data inicial")
            data_fim = self._validar_data(self.filtro_data_fim.get().strip(), "Data final")

            if data_inicio and data_fim:
                inicio_dt = datetime.strptime(data_inicio, "%d/%m/%Y")
                fim_dt = datetime.strptime(data_fim, "%d/%m/%Y")
                if inicio_dt > fim_dt:
                    raise ValueError("A data inicial não pode ser maior que a data final.")

            status = self.filtro_status.get().strip() if hasattr(self, "filtro_status") else "Todos"
            rendimentos = self.service.listar(data_inicio=data_inicio, data_fim=data_fim, status=status)

            if not rendimentos:
                messagebox.showinfo("Exportar Excel", "Não há rendimentos para exportar com os filtros atuais.")
                return

            arquivo = filedialog.asksaveasfilename(
                title="Salvar exportação de rendimentos",
                defaultextension=".xlsx",
                filetypes=[("Planilha Excel", "*.xlsx")],
                initialfile=f"rendimentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
            if not arquivo:
                return

            workbook = Workbook()
            sheet_rendimentos = workbook.active
            sheet_rendimentos.title = "Rendimentos"

            cabecalho = [
                "ID",
                "Cliente",
                "Valor Inicial (R$)",
                "Data Inicial",
                "Forma Inicial",
                "Status Inicial",
                "Valor Final (R$)",
                "Data Final",
                "Forma Final",
                "Status Final",
                "Responsável",
            ]
            sheet_rendimentos.append(cabecalho)

            for celula in sheet_rendimentos[1]:
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill(fill_type="solid", fgColor="A66850")
                celula.alignment = Alignment(horizontal="center", vertical="center")

            total_inicial = 0.0
            total_final = 0.0
            recebidos_completos = 0
            pendentes = 0

            for rendimento in rendimentos:
                cliente_nome = self.cliente_map.get(rendimento.cliente_id, "Cliente não encontrado")
                valor_inicial = float(rendimento.pag_inicial_valor or 0.0)
                valor_final = float(rendimento.pag_final_valor or 0.0)

                total_inicial += valor_inicial
                total_final += valor_final

                if rendimento.pag_inicial_status == "Recebido" and rendimento.pag_final_status == "Recebido":
                    recebidos_completos += 1
                else:
                    pendentes += 1

                sheet_rendimentos.append([
                    rendimento.id,
                    cliente_nome,
                    valor_inicial,
                    rendimento.pag_inicial_data or "",
                    rendimento.pag_inicial_forma or "",
                    rendimento.pag_inicial_status,
                    valor_final,
                    rendimento.pag_final_data or "",
                    rendimento.pag_final_forma or "",
                    rendimento.pag_final_status,
                    rendimento.responsavel or "",
                ])

            sheet_resumo = workbook.create_sheet("Resumo")
            sheet_resumo.append(["Métrica", "Valor"])

            for celula in sheet_resumo[1]:
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill(fill_type="solid", fgColor="565b5e")
                celula.alignment = Alignment(horizontal="center", vertical="center")

            sheet_resumo.append(["Quantidade de rendimentos", len(rendimentos)])
            sheet_resumo.append(["Total Inicial (R$)", total_inicial])
            sheet_resumo.append(["Total Final (R$)", total_final])
            sheet_resumo.append(["Total Geral (R$)", total_inicial + total_final])
            sheet_resumo.append(["Rendimentos recebidos", recebidos_completos])
            sheet_resumo.append(["Rendimentos pendentes", pendentes])

            for planilha in (sheet_rendimentos, sheet_resumo):
                larguras = {}
                for linha in planilha.iter_rows():
                    for celula in linha:
                        valor = "" if celula.value is None else str(celula.value)
                        larguras[celula.column_letter] = max(larguras.get(celula.column_letter, 0), len(valor))

                for coluna, largura in larguras.items():
                    planilha.column_dimensions[coluna].width = min(largura + 2, 45)

                planilha.freeze_panes = "A2"

            workbook.save(arquivo)
            messagebox.showinfo("Exportar Excel", f"Exportação concluída com sucesso.\nArquivo salvo em:\n{arquivo}")
        except ValueError as exc:
            messagebox.showerror("Erro", str(exc))
        except Exception as exc:
            messagebox.showerror("Exportar Excel", f"Não foi possível exportar os rendimentos.\n\n{exc}")

    def _limpar_form(self):
        self.current_rendimento_id = None
        if self.cb_cliente.cget("values"):
            self.cb_cliente.set(self.cb_cliente.cget("values")[0])
        self.entry_responsavel.delete(0, "end")

        self.entry_pag_ini_valor.delete(0, "end")
        self.entry_pag_ini_valor.insert(0, "0,00")
        self.entry_pag_ini_data.delete(0, "end")
        self.cb_pag_ini_forma.set("Pix")
        self.cb_pag_ini_status.set("Pendente")

        self.entry_pag_fin_valor.delete(0, "end")
        self.entry_pag_fin_valor.insert(0, "0,00")
        self.entry_pag_fin_data.delete(0, "end")
        self.cb_pag_fin_forma.set("Pix")
        self.cb_pag_fin_status.set("Pendente")

        self.btn_excluir.configure(state="disabled")

    def _on_novo(self):
        if not self.cliente_map:
            messagebox.showwarning("Aviso", "Cadastre ao menos um cliente antes de lançar um rendimento.")
            return
        self._limpar_form()
        self._show_form(editando=False)

    def _on_editar_selecionado(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return

        rendimento_id = int(self.tree.item(selected[0])["values"][0])
        rendimento = self.service.get_by_id(rendimento_id)
        if not rendimento:
            return

        self._limpar_form()
        self.current_rendimento_id = rendimento.id

        if rendimento.cliente_id in self.cliente_map:
            self.cb_cliente.set(self.cliente_map[rendimento.cliente_id])

        self.entry_responsavel.delete(0, "end")
        self.entry_responsavel.insert(0, rendimento.responsavel or "")

        self.entry_pag_ini_valor.delete(0, "end")
        self.entry_pag_ini_valor.insert(0, self._formatar_moeda_ui(rendimento.pag_inicial_valor or 0.0))
        self.entry_pag_ini_data.delete(0, "end")
        self.entry_pag_ini_data.insert(0, rendimento.pag_inicial_data or "")
        self.cb_pag_ini_forma.set(rendimento.pag_inicial_forma or "Pix")
        self.cb_pag_ini_status.set(rendimento.pag_inicial_status or "Pendente")

        self.entry_pag_fin_valor.delete(0, "end")
        self.entry_pag_fin_valor.insert(0, self._formatar_moeda_ui(rendimento.pag_final_valor or 0.0))
        self.entry_pag_fin_data.delete(0, "end")
        self.entry_pag_fin_data.insert(0, rendimento.pag_final_data or "")
        self.cb_pag_fin_forma.set(rendimento.pag_final_forma or "Pix")
        self.cb_pag_fin_status.set(rendimento.pag_final_status or "Pendente")

        self.btn_excluir.configure(state="normal")
        self._show_form(editando=True)

    def _on_salvar(self):
        try:
            if not self.cliente_map:
                raise ValueError("Cadastre ao menos um cliente antes de lançar um rendimento.")

            cliente_label = self.cb_cliente.get().strip()
            if " - " not in cliente_label:
                raise ValueError("Selecione um cliente válido.")
            cliente_id = int(cliente_label.split(" - ", 1)[0])

            pag_ini_valor = self._parse_float_campo(self.entry_pag_ini_valor.get().strip(), "Valor Inicial", obrigatorio=False, minimo=0.0)
            pag_ini_data = self._validar_data(self.entry_pag_ini_data.get().strip(), "Data Inicial")
            pag_ini_forma = self.cb_pag_ini_forma.get().strip() or None
            pag_ini_status = self.cb_pag_ini_status.get().strip()

            pag_fin_valor = self._parse_float_campo(self.entry_pag_fin_valor.get().strip(), "Valor Final", obrigatorio=False, minimo=0.0)
            pag_fin_data = self._validar_data(self.entry_pag_fin_data.get().strip(), "Data Final")
            pag_fin_forma = self.cb_pag_fin_forma.get().strip() or None
            pag_fin_status = self.cb_pag_fin_status.get().strip()

            if pag_ini_status not in ["Pendente", "Recebido"]:
                raise ValueError("Status Inicial inválido.")
            if pag_fin_status not in ["Pendente", "Recebido"]:
                raise ValueError("Status Final inválido.")

            if pag_ini_status == "Recebido" and not pag_ini_data:
                raise ValueError("Informe a data do pagamento inicial recebido.")
            if pag_fin_status == "Recebido" and not pag_fin_data:
                raise ValueError("Informe a data do pagamento final recebido.")

            rendimento = Rendimento(
                id=self.current_rendimento_id,
                cliente_id=cliente_id,
                pag_inicial_valor=pag_ini_valor,
                pag_inicial_data=pag_ini_data or None,
                pag_inicial_forma=pag_ini_forma,
                pag_inicial_status=pag_ini_status,
                pag_final_valor=pag_fin_valor,
                pag_final_data=pag_fin_data or None,
                pag_final_forma=pag_fin_forma,
                pag_final_status=pag_fin_status,
                responsavel=self.entry_responsavel.get().strip() or None,
            )

            self.service.salvar(rendimento)
            self._carregar_dados_seguro()
            self._show_lista()

        except ValueError as exc:
            messagebox.showerror("Erro", str(exc))

    def _on_excluir(self):
        if not self.current_rendimento_id:
            return

        cliente = self.cb_cliente.get().strip() or f"ID {self.current_rendimento_id}"
        mensagem = (
            f"Confirma a exclusao do rendimento '{cliente}'?\n\n"
            "Esta acao nao pode ser desfeita."
        )
        if messagebox.askyesno("Confirmar exclusao", mensagem):
            self.service.excluir(self.current_rendimento_id)
            self._carregar_dados_seguro()
            self._show_lista()

    def _aplicar_mascara_data(self, event):
        entry = event.widget
        digits = "".join(ch for ch in entry.get() if ch.isdigit())[:8]

        partes = []
        if len(digits) >= 2:
            partes.append(digits[:2])
        elif digits:
            partes.append(digits)

        if len(digits) >= 4:
            partes.append(digits[2:4])
        elif len(digits) > 2:
            partes.append(digits[2:])

        if len(digits) > 4:
            partes.append(digits[4:])

        formatado = "/".join(partes)
        entry.delete(0, "end")
        entry.insert(0, formatado)

    def _aplicar_mascara_moeda(self, event):
        entry = event.widget
        bruto = entry.get()
        digitos = "".join(ch for ch in bruto if ch.isdigit())
        tem_virgula = "," in bruto

        if not digitos:
            sanitizado = ""
        else:
            if tem_virgula:
                parte_int, parte_dec = bruto.split(",", 1)
                int_digits = "".join(ch for ch in parte_int if ch.isdigit()) or "0"
                dec_all = "".join(ch for ch in parte_dec if ch.isdigit())
                if len(dec_all) > 2:
                    int_digits = int_digits + dec_all[:-2]
                    dec_digits = dec_all[-2:]
                else:
                    dec_digits = dec_all.ljust(2, "0")
            else:
                int_digits = digitos
                dec_digits = "00"

            int_norm = str(int(int_digits))
            grupos = []
            while int_norm:
                grupos.insert(0, int_norm[-3:])
                int_norm = int_norm[:-3]

            sanitizado = f"{'.'.join(grupos)},{dec_digits}"

        entry.delete(0, "end")
        entry.insert(0, sanitizado)

    def _validar_data(self, valor: str, campo: str, obrigatorio: bool = False) -> str:
        if not valor:
            if obrigatorio:
                raise ValueError(f"O campo {campo} é obrigatório.")
            return ""

        try:
            datetime.strptime(valor, "%d/%m/%Y")
        except ValueError as exc:
            raise ValueError(f"O campo {campo} deve estar no formato DD/MM/AAAA.") from exc

        return valor

    def _parse_float_campo(self, valor: str, campo: str, obrigatorio: bool = False, minimo: float | None = None) -> float:
        numero_txt = self._normalizar_numero_para_float(valor)
        if not numero_txt:
            if obrigatorio:
                raise ValueError(f"O campo {campo} é obrigatório.")
            return 0.0

        try:
            numero = float(numero_txt)
        except ValueError as exc:
            raise ValueError(f"O campo {campo} deve ser numérico.") from exc

        if minimo is not None and numero < minimo:
            raise ValueError(f"O campo {campo} não pode ser menor que {minimo}.")

        return numero

    def _normalizar_numero_para_float(self, valor: str) -> str:
        numero_txt = valor.strip().replace(" ", "")
        if not numero_txt:
            return ""

        if "," in numero_txt:
            return numero_txt.replace(".", "").replace(",", ".")

        return numero_txt

    def _formatar_moeda_ui(self, valor: float) -> str:
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
