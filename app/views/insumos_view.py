"""
InsumosView — Insumos e Histórico de Preços em uma única view com abas.
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Callable, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.formatters import fmt_data, fmt_moeda, parse_float
from app.core import event_bus
from app.models.insumo import Insumo
from app.services.insumo_service import InsumoService
from app.core.enums import UnidadeMedida
from app.ui.theme import (
    ACCENT,
    BG_DEEP,
    CARD_BG,
    CARD_BORDER,
    COLOR_GREEN,
    FIELD_BG,
    HEADER_BG,
    TEXT_MUTED,
    TEXT_PRIMARY,
    TEXT_SECONDARY,
    _btn_accent,
    _btn_danger,
    _btn_ghost,
    _card,
    _combo,
    _entry,
    _optmenu,
    _sep,
    _treeview_style,
)


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  PAINEL DE INSUMOS                                                          ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class InsumosPanel(ctk.CTkFrame):
    """Lista + formulário de insumos. Sem topbar própria."""

    def __init__(self, master,
                 on_state_change: Optional[Callable[[bool], None]] = None,
                 on_estoque_alerta_change: Optional[Callable[[int], None]] = None,
                 **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.service = InsumoService()
        self.current_insumo_id = None
        self._on_state = on_state_change
        self._on_estoque_alerta = on_estoque_alerta_change
        self._editing = False

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_lista()
        self._build_form()
        self._show_lista()
        self._carregar_dados()

    # ── estado público ────────────────────────────────────────────────────
    def is_editing(self) -> bool:
        return self._editing

    def go_new(self):
        self.current_insumo_id = None
        self.entry_nome.delete(0, "end")
        self.combo_categoria.set("Ingrediente")
        self.entry_peso.delete(0, "end")
        self.combo_medida.set(UnidadeMedida.G.value)
        self.entry_preco.delete(0, "end")
        self.entry_qtd.delete(0, "end"); self.entry_qtd.insert(0, "0")
        self.entry_qtd_min.delete(0, "end"); self.entry_qtd_min.insert(0, "0")
        self.lbl_custo_calc.configure(text="Custo/Un: R$ 0.00")
        self.btn_excluir.configure(state="disabled")
        self._show_form(editando=False)

    def export_excel(self):
        self._on_exportar_excel()
        
    def refresh(self):
        """Atualiza a lista de insumos."""
        self._carregar_dados()

    # ── lista ─────────────────────────────────────────────────────────────
    def _build_lista(self):
        self._frame_lista = ctk.CTkFrame(self, fg_color="transparent")
        self._frame_lista.grid_columnconfigure(0, weight=1)
        self._frame_lista.grid_rowconfigure(1, weight=1)

        # filtros
        fc = _card(self._frame_lista)
        fc.grid(row=0, column=0, padx=4, pady=(0, 12), sticky="ew")
        fc.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fc, text="FILTROS", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, padx=16, pady=(12, 6), sticky="w")

        ctrl = ctk.CTkFrame(fc, fg_color="transparent")
        ctrl.grid(row=1, column=0, padx=16, pady=(0, 12), sticky="ew")

        self.entry_busca = _entry(ctrl, placeholder_text="Buscar por nome...", width=320)
        self.entry_busca.pack(side="left", padx=(0, 10))
        self.entry_busca.bind("<KeyRelease>", lambda e: self._carregar_dados())

        self.combo_categoria_filtro = _combo(
            ctrl, values=["Todos", "Ingrediente", "Embalagem", "Gás"],
            command=lambda e: self._carregar_dados(), width=180,
        )
        self.combo_categoria_filtro.pack(side="left", padx=(0, 10))
        self.combo_categoria_filtro.set("Todos")

        _btn_accent(ctrl, "Buscar", self._carregar_dados).pack(side="left", padx=(0, 8))
        _btn_ghost(ctrl, "Limpar", self._limpar_filtros).pack(side="left")

        # tabela
        tf = _card(self._frame_lista)
        tf.grid(row=1, column=0, sticky="nsew", padx=4)
        tf.grid_columnconfigure(0, weight=1); tf.grid_rowconfigure(0, weight=1)

        style = _treeview_style("Insumos", rowheight=32)
        self.tree = ttk.Treeview(tf,
                                 columns=("id", "nome", "categoria", "peso", "preco", "custo", "qtd"),
                                 show="headings", style=style)
        for col, txt, w, anc in [
            ("id",        "ID",             40, "center"),
            ("nome",      "Nome",          160, "w"),
            ("categoria", "Categoria",     110, "center"),
            ("peso",      "Peso/Vol",       90, "center"),
            ("preco",     "Preço (R$)",     90, "center"),
            ("custo",     "Custo/Un (R$)", 110, "center"),
            ("qtd",       "Qtd Disp.",      90, "center"),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor=anc)

        # alertas com tons suaves Dolce
        self.tree.tag_configure("estoque_alerta",  background="#FFF4CC", foreground="#7A5500")
        self.tree.tag_configure("estoque_critico", background="#FAE8EC", foreground="#A0404E")

        self.tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        sb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        sb.grid(row=0, column=1, sticky="ns", padx=(0, 2), pady=2)
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.bind("<Double-1>",  self._on_double_click)
        self.tree.bind("<Button-3>",  self._show_context_menu)

        self.context_menu = tk.Menu(self, tearoff=0,
                                    bg=CARD_BG, fg=TEXT_PRIMARY,
                                    activebackground="#FAE8EC", activeforeground=ACCENT)
        self.context_menu.add_command(label="Editar",         command=self._on_editar_selecionado)
        self.context_menu.add_command(label="Lançar Compra",  command=self._on_compra_selecionada)
        self.context_menu.add_command(label="Excluir",        command=self._on_excluir_selecionado)

    # ── formulário ────────────────────────────────────────────────────────
    def _build_form(self):
        self._frame_form = ctk.CTkFrame(self, fg_color="transparent")
        self._frame_compra = ctk.CTkFrame(self, fg_color="transparent")

        # --- FORM CADASTRO ---
        # Atalhos de teclado
        self._frame_form.bind("<Control-s>", lambda e: self._on_salvar())
        self._frame_form.bind("<Escape>",    lambda e: self._ocultar_form())

        fc = _card(self._frame_form)
        fc.pack(fill="both", expand=True, padx=4, pady=0)

        inner = ctk.CTkFrame(fc, fg_color="transparent")
        inner.pack(expand=True, fill="both", padx=18, pady=14)
        inner.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(inner, text="DADOS DO INSUMO", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="w")

        ctk.CTkLabel(inner, text="Nome*", text_color=TEXT_SECONDARY
                     ).grid(row=1, column=0, padx=12, pady=(5, 2), sticky="w")
        self.entry_nome = _entry(inner, placeholder_text="Texto")
        self.entry_nome.grid(row=2, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(inner, text="Categoria*", text_color=TEXT_SECONDARY
                     ).grid(row=3, column=0, padx=12, pady=(5, 2), sticky="w")
        self.combo_categoria = _combo(inner, values=["Ingrediente", "Embalagem", "Gás"])
        self.combo_categoria.grid(row=4, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(inner, text="Peso/Volume Total*", text_color=TEXT_SECONDARY
                     ).grid(row=5, column=0, padx=12, pady=(5, 2), sticky="w")
        ctk.CTkLabel(inner, text="Unidade*", text_color=TEXT_SECONDARY
                     ).grid(row=5, column=1, padx=12, pady=(5, 2), sticky="w")
        self.entry_peso = _entry(inner, placeholder_text="Numero")
        self.entry_peso.grid(row=6, column=0, padx=12, pady=(0, 10), sticky="ew")
        self.combo_medida = _combo(inner, values=[UnidadeMedida.G.value, UnidadeMedida.ML.value, UnidadeMedida.UNIDADE.value], width=120)
        self.combo_medida.grid(row=6, column=1, padx=12, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(inner, text="Preço de Compra Atual (R$)*", text_color=TEXT_SECONDARY
                     ).grid(row=7, column=0, padx=12, pady=(5, 2), sticky="w")
        self.entry_preco = _entry(inner, placeholder_text="0,00")
        self.entry_preco.grid(row=8, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="ew")

        self.lbl_custo_calc = ctk.CTkLabel(inner, text="Custo/Un: R$ 0.00",
                                           text_color=ACCENT,
                                           font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_custo_calc.grid(row=9, column=0, columnspan=2, padx=12, pady=6, sticky="w")

        self.entry_peso.bind("<KeyRelease>",
                             lambda e: (self._aplicar_mascara_decimal(e), self._atualizar_custo_label()))
        self.entry_preco.bind("<KeyRelease>", self._atualizar_custo_label)

        ctk.CTkLabel(inner, text="Quantidade Disponível", text_color=TEXT_SECONDARY
                     ).grid(row=10, column=0, padx=12, pady=(5, 2), sticky="w")
        ctk.CTkLabel(inner, text="Quantidade Mínima", text_color=TEXT_SECONDARY
                     ).grid(row=10, column=1, padx=12, pady=(5, 2), sticky="w")
        self.entry_qtd = _entry(inner, placeholder_text="Numero")
        self.entry_qtd.grid(row=11, column=0, padx=12, pady=(0, 10), sticky="ew")
        self.entry_qtd.bind("<KeyRelease>", self._aplicar_mascara_decimal)
        self.entry_qtd_min = _entry(inner, placeholder_text="Numero")
        self.entry_qtd_min.grid(row=11, column=1, padx=12, pady=(0, 10), sticky="ew")
        self.entry_qtd_min.bind("<KeyRelease>", self._aplicar_mascara_decimal)

        fa = ctk.CTkFrame(inner, fg_color="transparent")
        fa.grid(row=12, column=0, columnspan=2, padx=12, pady=(14, 8), sticky="e")
        self.btn_salvar   = _btn_accent(fa, "Salvar",   self._on_salvar)
        self.btn_excluir  = _btn_danger(fa, "Excluir",  self._excluir_form)
        self.btn_cancelar = _btn_ghost(fa, "Cancelar",  self._ocultar_form)
        for b in (self.btn_cancelar, self.btn_excluir, self.btn_salvar):
            b.pack(side="right", padx=5)

        ctk.CTkLabel(inner, text="* Campos obrigatórios", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10)).grid(row=13, column=0, columnspan=2, padx=12, pady=(0, 5), sticky="w")

        # --- FORM COMPRA ---
        self._frame_compra.bind("<Escape>", lambda e: self._ocultar_form())
        
        fcomp = _card(self._frame_compra)
        fcomp.pack(fill="both", expand=True, padx=4, pady=0)
        
        icomp = ctk.CTkFrame(fcomp, fg_color="transparent")
        icomp.pack(expand=True, fill="both", padx=18, pady=14)
        icomp.grid_columnconfigure((0, 1), weight=1)
        
        ctk.CTkLabel(icomp, text="LANÇAR COMPRA DE INSUMO", text_color=ACCENT,
                     font=ctk.CTkFont(size=12, weight="bold")
                     ).grid(row=0, column=0, columnspan=2, padx=12, pady=(0, 10), sticky="w")
        
        self.lbl_compra_insumo = ctk.CTkLabel(icomp, text="Insumo: ...", text_color=TEXT_PRIMARY,
                                              font=ctk.CTkFont(size=14, weight="bold"))
        self.lbl_compra_insumo.grid(row=1, column=0, columnspan=2, padx=12, pady=(0, 15), sticky="w")

        # Dados da mercadoria
        ctk.CTkLabel(icomp, text="Quantidade Comprada", text_color=TEXT_SECONDARY
                     ).grid(row=2, column=0, padx=12, pady=(5, 2), sticky="w")
        self.entry_compra_qtd = _entry(icomp, placeholder_text="0,00")
        self.entry_compra_qtd.grid(row=3, column=0, padx=12, pady=(0, 10), sticky="ew")
        self.entry_compra_qtd.bind("<KeyRelease>", self._aplicar_mascara_decimal)

        ctk.CTkLabel(icomp, text="Valor Total da Nota (R$)*", text_color=TEXT_SECONDARY
                     ).grid(row=2, column=1, padx=12, pady=(5, 2), sticky="w")
        self.entry_compra_valor = _entry(icomp, placeholder_text="0,00")
        self.entry_compra_valor.grid(row=3, column=1, padx=12, pady=(0, 10), sticky="ew")
        self.entry_compra_valor.bind("<KeyRelease>", self._aplicar_mascara_decimal)

        # Dados do pagamento
        ctk.CTkLabel(icomp, text="Data da Compra*", text_color=TEXT_SECONDARY
                     ).grid(row=4, column=0, padx=12, pady=(5, 2), sticky="w")
        self.entry_compra_data = _entry(icomp, placeholder_text="DD/MM/AAAA")
        self.entry_compra_data.grid(row=5, column=0, padx=12, pady=(0, 10), sticky="ew")
        self.entry_compra_data.bind("<KeyRelease>", self._aplicar_mascara_data)

        ctk.CTkLabel(icomp, text="Forma de Pagamento", text_color=TEXT_SECONDARY
                     ).grid(row=4, column=1, padx=12, pady=(5, 2), sticky="w")
        self.cb_compra_forma = _optmenu(icomp, ["PIX", "Dinheiro", "Cartão Crédito", "Cartão Débito", "Boleto", "Transferência"])
        self.cb_compra_forma.grid(row=5, column=1, padx=12, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(icomp, text="Status Pagamento", text_color=TEXT_SECONDARY
                     ).grid(row=6, column=0, padx=12, pady=(5, 2), sticky="w")
        self.cb_compra_status = _optmenu(icomp, ["Pago", "Pendente"])
        self.cb_compra_status.grid(row=7, column=0, padx=12, pady=(0, 10), sticky="ew")
        
        ctk.CTkLabel(icomp, text="Responsável", text_color=TEXT_SECONDARY
                     ).grid(row=6, column=1, padx=12, pady=(5, 2), sticky="w")
        self.entry_compra_resp = _entry(icomp, placeholder_text="Nome")
        self.entry_compra_resp.grid(row=7, column=1, padx=12, pady=(0, 10), sticky="ew")

        fca = ctk.CTkFrame(icomp, fg_color="transparent")
        fca.grid(row=8, column=0, columnspan=2, padx=12, pady=(20, 8), sticky="e")
        self.btn_compra_salvar = _btn_accent(fca, "Confirmar Compra", self._on_salvar_compra)
        self.btn_compra_cancel = _btn_ghost(fca, "Cancelar", self._ocultar_form)
        self.btn_compra_cancel.pack(side="right", padx=5)
        self.btn_compra_salvar.pack(side="right", padx=5)

    # ── transições ────────────────────────────────────────────────────────
    def _show_lista(self):
        self._editing = False
        self._frame_form.grid_forget()
        self._frame_compra.grid_forget()
        self._frame_lista.grid(row=0, column=0, sticky="nsew")
        self._frame_lista.grid_columnconfigure(0, weight=1)
        self._frame_lista.grid_rowconfigure(1, weight=1)
        if self._on_state: self._on_state(False)

    def _show_form(self, editando=False):
        self._editing = True
        self._frame_lista.grid_forget()
        self._frame_compra.grid_forget()
        self._frame_form.grid(row=0, column=0, sticky="nsew")
        self._frame_form.grid_columnconfigure(0, weight=1)
        self._frame_form.grid_rowconfigure(0, weight=1)
        if self._on_state: self._on_state(True)
        self.entry_nome.focus_set()

    def _show_compra_form(self):
        self._editing = True
        self._frame_lista.grid_forget()
        self._frame_form.grid_forget()
        self._frame_compra.grid(row=0, column=0, sticky="nsew")
        self._frame_compra.grid_columnconfigure(0, weight=1)
        self._frame_compra.grid_rowconfigure(0, weight=1)
        if self._on_state: self._on_state(True)
        self.entry_compra_qtd.focus_set()

    def _ocultar_form(self):
        self._show_lista()

    # ── dados ─────────────────────────────────────────────────────────────
    def _limpar_filtros(self):
        self.entry_busca.delete(0, "end")
        self.combo_categoria_filtro.set("Todos")
        self._carregar_dados()

    def _atualizar_custo_label(self, event=None):
        try:
            peso = parse_float(self.entry_peso.get(), "Peso/Volume Total", obrigatorio=False)
            preco = parse_float(self.entry_preco.get(), "Preço de Compra", obrigatorio=False)
            self.lbl_custo_calc.configure(
                text=f"Custo/Un: R$ {preco/peso:.4f}" if peso > 0 else "Custo/Un: R$ 0.00")
        except ValueError:
            self.lbl_custo_calc.configure(text="Custo/Un: Valores inválidos")

    def _carregar_dados(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        insumos = self.service.listar(
            nome=self.entry_busca.get(),
            categoria=self.combo_categoria_filtro.get(),
        )
        for i in insumos:
            tag = ()
            if i.quantidade_disponivel <= 0:
                tag = ("estoque_critico",)
            elif i.quantidade_disponivel <= i.quantidade_minima:
                tag = ("estoque_alerta",)
            self.tree.insert("", "end", values=(
                i.id, i.nome, i.categoria,
                f"{i.peso_volume_total} {i.unidade_medida}",
                fmt_moeda(i.preco_compra, 2),
                fmt_moeda(i.custo_por_unidade, 4),
                i.quantidade_disponivel,
            ), tags=tag)
        
        # Regra I-05: callback on_estoque_alerta_change recebe o total de itens com estoque baixo
        if callable(self._on_estoque_alerta):
            total_baixo = sum(1 for i in insumos if i.quantidade_disponivel <= i.quantidade_minima)
            self._on_estoque_alerta(total_baixo)

    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self._on_editar_selecionado()

    def _show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.tk_popup(event.x_root, event.y_root)

    def _on_editar_selecionado(self):
        sel = self.tree.selection()
        if not sel: return
        insumo = self.service.get_by_id(self.tree.item(sel[0])["values"][0])
        if not insumo: return
        self.current_insumo_id = insumo.id
        self.entry_nome.delete(0, "end"); self.entry_nome.insert(0, insumo.nome)
        self.combo_categoria.set(insumo.categoria)
        self.entry_peso.delete(0, "end"); self.entry_peso.insert(0, str(insumo.peso_volume_total))
        self.combo_medida.set(insumo.unidade_medida)
        self.entry_preco.delete(0, "end"); self.entry_preco.insert(0, fmt_moeda(insumo.preco_compra, 2))
        self.entry_qtd.delete(0, "end"); self.entry_qtd.insert(0, str(insumo.quantidade_disponivel))
        self.entry_qtd_min.delete(0, "end"); self.entry_qtd_min.insert(0, str(insumo.quantidade_minima))
        self._atualizar_custo_label()
        self.btn_excluir.configure(state="normal")
        self._show_form(editando=True)

    def _on_compra_selecionada(self):
        sel = self.tree.selection()
        if not sel: return
        insumo = self.service.get_by_id(self.tree.item(sel[0])["values"][0])
        if not insumo: return
        self.current_insumo_id = insumo.id
        self.lbl_compra_insumo.configure(text=f"Insumo: {insumo.nome} ({insumo.unidade_medida})")
        
        # Reset compra form
        self.entry_compra_qtd.delete(0, "end");   self.entry_compra_qtd.insert(0, str(insumo.peso_volume_total))
        self.entry_compra_valor.delete(0, "end"); self.entry_compra_valor.insert(0, fmt_moeda(insumo.preco_compra, 2))
        self.entry_compra_data.delete(0, "end");  self.entry_compra_data.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.cb_compra_forma.set("PIX")
        self.cb_compra_status.set("Pago")
        self.entry_compra_resp.delete(0, "end")
        
        self._show_compra_form()

    def _excluir_form(self):
        if not self.current_insumo_id: return
        nome = self.entry_nome.get()
        if messagebox.askyesno("Confirmar exclusão",
                               f"Excluir o insumo '{nome}'?\nEsta ação não pode ser desfeita."):
            try:
                self.service.excluir(self.current_insumo_id)
                self._carregar_dados(); self._ocultar_form()
            except ValueError as e:
                messagebox.showerror("Erro de Vínculo", str(e))

    def _on_excluir_selecionado(self):
        sel = self.tree.selection()
        if not sel: return
        item_id = self.tree.item(sel[0])["values"][0]
        nome    = self.tree.item(sel[0])["values"][1]
        if messagebox.askyesno("Confirmar exclusão",
                               f"Excluir o insumo '{nome}'?\nEsta ação não pode ser desfeita."):
            try:
                self.service.excluir(item_id)
                self._carregar_dados()
                if self.current_insumo_id == item_id:
                    self._ocultar_form()
            except ValueError as e:
                messagebox.showerror("Erro de Vínculo", str(e))

    def _on_salvar(self):
        try:
            nome      = self.entry_nome.get().strip()
            categoria = self.combo_categoria.get().strip()
            medida    = self.combo_medida.get().strip()
            if not nome: raise ValueError("Nome é obrigatório.")
            if categoria not in ["Ingrediente", "Embalagem", "Gás"]: raise ValueError("Categoria inválida.")
            if medida    not in [UnidadeMedida.G.value, UnidadeMedida.ML.value, UnidadeMedida.UNIDADE.value]:              raise ValueError("Unidade inválida.")
            peso    = parse_float(self.entry_peso.get(), "Peso/Volume Total", minimo=0.000001)
            preco   = parse_float(self.entry_preco.get(), "Preço de Compra", minimo=0.0)
            qtd     = parse_float(self.entry_qtd.get(), "Quantidade Disponível", obrigatorio=False, minimo=0.0)
            qtd_min = parse_float(self.entry_qtd_min.get(), "Quantidade Mínima", obrigatorio=False, minimo=0.0)
            if qtd_min > qtd:
                raise ValueError("Quantidade mínima não pode ser maior que a disponível.")
            self.service.salvar(Insumo(
                id=self.current_insumo_id, nome=nome, categoria=categoria,
                peso_volume_total=peso, unidade_medida=medida, preco_compra=preco,
                quantidade_disponivel=qtd, quantidade_minima=qtd_min,
            ))
            self._ocultar_form(); self._carregar_dados()
        except ValueError as e:
            messagebox.showerror("Erro", str(e))

    def _on_salvar_compra(self):
        try:
            if not self.current_insumo_id: return
            insumo = self.service.get_by_id(self.current_insumo_id)
            if not insumo: return
            
            qtd_compra   = parse_float(self.entry_compra_qtd.get(), "Quantidade Comprada", minimo=0.000001)
            valor_total  = parse_float(self.entry_compra_valor.get(), "Valor Total", minimo=0.01)
            data_compra  = parse_data(self.entry_compra_data.get(), "Data da Compra", obrigatorio=True)
            forma_pag    = self.cb_compra_forma.get()
            status_pag   = self.cb_compra_status.get()
            responsavel  = self.entry_compra_resp.get().strip() or None
            
            # Atualiza o insumo: soma estoque e define o novo preço de compra unitário
            novo_preco_unidade = valor_total / qtd_compra
            novo_preco_cadastro = novo_preco_unidade * insumo.peso_volume_total
            
            insumo.quantidade_disponivel += qtd_compra
            insumo.preco_compra = novo_preco_cadastro
            # Regra I-06: define data_compra automática na compra
            insumo.data_compra = data_compra
            
            self.service.salvar(insumo)
            
            # Emite evento para gerar Despesa
            event_bus.emit(
                "insumo.comprado",
                insumo_id=insumo.id,
                valor_total=valor_total,
                data_pagamento=data_compra,
                forma_pagamento=forma_pag,
                status_pagamento=status_pag,
                responsavel=responsavel
            )
            
            messagebox.showinfo("Sucesso", "Compra registrada com sucesso!\nEstoque atualizado e despesa gerada.")
            self._ocultar_form(); self._carregar_dados()
            
        except ValueError as e:
            messagebox.showerror("Erro", str(e))

    def _on_exportar_excel(self):
        insumos = self.service.listar(
            nome=self.entry_busca.get().strip(),
            categoria=self.combo_categoria_filtro.get().strip(),
        )
        if not insumos:
            messagebox.showinfo("Exportar Excel", "Não há insumos para exportar."); return
        arq = filedialog.asksaveasfilename(
            title="Salvar exportação de insumos", defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            initialfile=f"insumos_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        )
        if not arq: return
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Insumos"
            hdr = ["ID","Nome","Categoria","Peso/Volume Total","Unidade",
                   "Preço de Compra","Custo por Unidade","Qtd Disponível","Qtd Mínima","Data Compra"]
            ws.append(hdr)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C96B7A")
                c.alignment = Alignment(horizontal="center", vertical="center")
            for i in insumos:
                dc = i.data_compra
                if dc:
                    try: dc = datetime.strptime(dc, "%Y-%m-%d").strftime("%d/%m/%Y")
                    except ValueError: pass
                ws.append([i.id, i.nome, i.categoria, i.peso_volume_total, i.unidade_medida,
                           i.preco_compra, i.custo_por_unidade,
                           i.quantidade_disponivel, i.quantidade_minima, dc or ""])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 2, 35)
            ws.freeze_panes = "A2"; wb.save(arq)
            messagebox.showinfo("Exportar Excel", f"Arquivo salvo em:\n{arq}")
        except Exception as e:
            messagebox.showerror("Exportar Excel", str(e))

    # ── utilitários ───────────────────────────────────────────────────────
    def _aplicar_mascara_data(self, event):
        entry  = event.widget
        digits = "".join(ch for ch in entry.get() if ch.isdigit())[:8]
        partes = []
        if len(digits) >= 2: partes.append(digits[:2])
        elif digits:          partes.append(digits)
        if len(digits) >= 4:  partes.append(digits[2:4])
        elif len(digits) > 2: partes.append(digits[2:])
        if len(digits) > 4:   partes.append(digits[4:])
        entry.delete(0, "end"); entry.insert(0, "/".join(partes))

    def _aplicar_mascara_decimal(self, event):
        e   = event.widget
        raw = e.get().replace(".", ",")
        res, tem_virgula = [], False
        for ch in raw:
            if ch.isdigit():                     res.append(ch)
            elif ch == "," and not tem_virgula:  res.append(ch); tem_virgula = True
        s = "".join(res)
        if "," in s:
            i, d = s.split(",", 1); s = f"{i},{d[:4]}"
        e.delete(0, "end"); e.insert(0, s)


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  PAINEL DE HISTÓRICO DE PREÇOS                                              ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class HistoricoPrecoPanel(ctk.CTkFrame):
    """Histórico de preços de insumos."""

    def __init__(self, master, service: Optional[InsumoService] = None, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.service     = service or InsumoService()
        self.insumo_map: dict[int, str] = {}

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self._build_filtros()
        self._build_conteudo()
        self._carregar_filtro_insumos()
        self._carregar_historico()

    def is_editing(self) -> bool:
        return False

    def go_new(self):
        pass

    def export_excel(self):
        pass

    def refresh(self):
        self._carregar_filtro_insumos()
        self._carregar_historico()

    # ── filtros ───────────────────────────────────────────────────────────
    def _build_filtros(self):
        fc = _card(self)
        fc.grid(row=0, column=0, padx=4, pady=(0, 10), sticky="ew")
        fc.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fc, text="FILTROS", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, padx=16, pady=(12, 6), sticky="w")

        ctrl = ctk.CTkFrame(fc, fg_color="transparent")
        ctrl.grid(row=1, column=0, padx=16, pady=(0, 12), sticky="ew")

        ctk.CTkLabel(ctrl, text="Insumo:", text_color=TEXT_SECONDARY).pack(side="left", padx=(0, 6))
        self.combo_insumo = _combo(ctrl, values=["Todos"], width=360)
        self.combo_insumo.pack(side="left", padx=(0, 10))
        self.combo_insumo.set("Todos")
        _btn_accent(ctrl, "Aplicar", self._carregar_historico, width=100).pack(side="left")

    # ── conteúdo ──────────────────────────────────────────────────────────
    def _build_conteudo(self):
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.grid(row=1, column=0, sticky="nsew", padx=4)
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)

        tf = _card(body)
        tf.grid(row=0, column=0, padx=(0, 6), sticky="nsew")
        tf.grid_columnconfigure(0, weight=1); tf.grid_rowconfigure(0, weight=1)

        style = _treeview_style("HistInsumo", rowheight=32)
        cols  = ("data", "insumo", "preco_anterior", "preco_novo", "variacao", "observacao")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", style=style)
        for col, txt, w, anc in [
            ("data",           "Data/Hora",      150, "center"),
            ("insumo",         "Insumo",         200, "w"),
            ("preco_anterior", "Preço Anterior",  120, "e"),
            ("preco_novo",     "Preço Novo",      120, "e"),
            ("variacao",       "Variação",         90, "center"),
            ("observacao",     "Observação",      230, "w"),
        ]:
            self.tree.heading(col, text=txt)
            self.tree.column(col, width=w, anchor=anc)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        sb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        sb.grid(row=0, column=1, sticky="ns", padx=(0, 2), pady=2)
        self.tree.configure(yscrollcommand=sb.set)

        tl = _card(body)
        tl.grid(row=0, column=1, padx=(6, 0), sticky="nsew")
        tl.grid_rowconfigure(1, weight=1); tl.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(tl, text="LINHA DO TEMPO", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, padx=16, pady=(12, 8), sticky="w")

        self.timeline_text = ctk.CTkTextbox(
            tl, wrap="word", fg_color=FIELD_BG,
            border_color=CARD_BORDER, text_color=TEXT_PRIMARY,
        )
        self.timeline_text.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="nsew")
        self.timeline_text.configure(state="disabled")

    # ── dados ─────────────────────────────────────────────────────────────
    def _carregar_filtro_insumos(self):
        self.insumo_map.clear()
        valores = ["Todos"]
        for i in self.service.listar():
            if i.id is None: continue
            lbl = f"{i.id} - {i.nome}"
            self.insumo_map[int(i.id)] = lbl
            valores.append(lbl)
        self.combo_insumo.configure(values=valores)
        self.combo_insumo.set("Todos")

    def _carregar_historico(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        insumo_id = self._insumo_id_filtro()
        historico = self.service.listar_historico_preco(insumo_id=insumo_id)
        for reg in historico:
            ant = float(reg.get("preco_anterior") or 0)
            nov = float(reg.get("preco_novo")     or 0)
            self.tree.insert("", "end", values=(
                fmt_data(reg.get("data_alteracao", "")),
                reg.get("insumo_nome", ""),
                f"R$ {fmt_moeda(ant)}",
                f"R$ {fmt_moeda(nov)}",
                self._fmt_variacao(ant, nov),
                reg.get("observacao") or "",
            ))
        self._render_timeline(historico)

    def _render_timeline(self, historico: list[dict]):
        linhas = []
        for reg in reversed(historico):
            ant = float(reg.get("preco_anterior") or 0)
            nov = float(reg.get("preco_novo")     or 0)
            linhas.append(
                f"{fmt_data(reg.get('data_alteracao',''))} | "
                f"{reg.get('insumo_nome','Insumo')}: "
                f"R$ {fmt_moeda(ant)} → R$ {fmt_moeda(nov)} "
                f"({self._fmt_variacao(ant, nov)})"
            )
        if not linhas:
            linhas = ["Sem alterações de preço para o filtro selecionado."]
        self.timeline_text.configure(state="normal")
        self.timeline_text.delete("1.0", "end")
        self.timeline_text.insert("1.0", "\n\n".join(linhas))
        self.timeline_text.configure(state="disabled")

    def _insumo_id_filtro(self) -> Optional[int]:
        s = self.combo_insumo.get().strip()
        if s == "Todos" or " - " not in s: return None
        return int(s.split(" - ", 1)[0])

    @staticmethod
    def _fmt_variacao(anterior: float, novo: float) -> str:
        if anterior == 0: return "0,00%" if novo == 0 else "Novo"
        return f"{((novo - anterior) / anterior) * 100:+.2f}%".replace(".", ",")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  INSUMOS VIEW  — container com abas                                         ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class InsumosView(ctk.CTkFrame):

    ABAS = ["Insumos", "Histórico de Preços"]

    def __init__(self, master,
                 on_estoque_alerta_change: Optional[Callable[[int], None]] = None,
                 **kw):
        super().__init__(master, fg_color=BG_DEEP, **kw)
        self._on_estoque_alerta = on_estoque_alerta_change
        self._aba_ativa = "Insumos"
        self._tab_btns: dict[str, tk.Frame]     = {}
        self._tab_lbls: dict[str, ctk.CTkLabel] = {}

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self._build_topbar()
        self._build_tabbar()
        self._build_content()
        self._switch_aba("Insumos")

    def refresh(self):
        self._panel_insumos._carregar_dados()
        self._panel_historico.refresh()

    # ── topbar ────────────────────────────────────────────────────────────
    def _build_topbar(self):
        bar = ctk.CTkFrame(self, fg_color=HEADER_BG, corner_radius=0, height=56)
        bar.grid(row=0, column=0, sticky="ew")
        bar.grid_propagate(False)
        bar.grid_columnconfigure(1, weight=1)

        self.lbl_title = ctk.CTkLabel(
            bar, text="Insumos",
            text_color=TEXT_PRIMARY,
            font=ctk.CTkFont(family="Roboto", size=20, weight="bold"),
        )
        self.lbl_title.grid(row=0, column=0, padx=20, sticky="w")

        self._action_bar = ctk.CTkFrame(bar, fg_color="transparent")
        self._action_bar.grid(row=0, column=1, padx=(6, 16), sticky="e")

        self.btn_novo   = _btn_accent(self._action_bar, "+ Novo Insumo", self._on_novo)
        self.btn_novo.pack(side="left", padx=4)

        self.btn_posicao = _btn_ghost(self._action_bar, "Posição de Estoque", self._on_posicao_estoque)
        self.btn_posicao.pack(side="left", padx=4)

        self.btn_export = _btn_ghost(self._action_bar, "Exportar Excel", self._on_export)
        self.btn_export.pack(side="left", padx=4)

    # ── tabbar ────────────────────────────────────────────────────────────
    def _build_tabbar(self):
        self._tabbar = ctk.CTkFrame(self, fg_color=HEADER_BG, corner_radius=0, height=40)
        self._tabbar.grid(row=1, column=0, sticky="ew")
        self._tabbar.grid_propagate(False)

        for aba in self.ABAS:
            tab = tk.Frame(self._tabbar, bg=HEADER_BG, cursor="hand2")
            tab.pack(side="left", padx=(16, 0))
            self._tab_btns[aba] = tab

            lbl = ctk.CTkLabel(tab, text=aba, fg_color="transparent",
                               text_color=TEXT_MUTED, font=ctk.CTkFont(size=13))
            lbl.pack(pady=(8, 0))
            self._tab_lbls[aba] = lbl

            ind = tk.Frame(tab, bg=HEADER_BG, height=2)
            ind.pack(fill="x", pady=(4, 0))
            tab._indicator = ind  # type: ignore

            tab.bind("<Button-1>", lambda e, a=aba: self._switch_aba(a))
            lbl.bind("<Button-1>", lambda e, a=aba: self._switch_aba(a))

        _sep(self).grid(row=1, column=0, sticky="sew")

    # ── área de conteúdo ──────────────────────────────────────────────────
    def _build_content(self):
        self._content = ctk.CTkFrame(self, fg_color="transparent")
        self._content.grid(row=2, column=0, sticky="nsew", padx=14, pady=(10, 14))
        self._content.grid_columnconfigure(0, weight=1)
        self._content.grid_rowconfigure(0, weight=1)

        shared_service = InsumoService()

        self._panel_insumos = InsumosPanel(
            self._content,
            on_state_change=self._on_panel_state,
            on_estoque_alerta_change=self._on_estoque_alerta,
        )
        self._panel_historico = HistoricoPrecoPanel(
            self._content,
            service=shared_service,
        )

    # ── lógica de abas ────────────────────────────────────────────────────
    def _switch_aba(self, aba: str):
        self._aba_ativa = aba

        for nome, tab in self._tab_btns.items():
            ativo = nome == aba
            lbl   = self._tab_lbls[nome]
            lbl.configure(
                text_color=TEXT_PRIMARY if ativo else TEXT_MUTED,
                font=ctk.CTkFont(size=13, weight="bold" if ativo else "normal"),
            )
            tab._indicator.configure(bg=ACCENT if ativo else HEADER_BG)

        self.lbl_title.configure(text=f"Insumos — {aba}")

        self._panel_insumos.grid_forget()
        self._panel_historico.grid_forget()

        panel = self._active_panel()
        if panel:
            panel.grid(row=0, column=0, sticky="nsew")
            if hasattr(panel, "refresh"):
                panel.refresh()

        if aba == "Insumos":
            self.btn_novo.pack(side="left", padx=4)
            self.btn_export.pack(side="left", padx=4)
        else:
            self.btn_novo.pack_forget()
            self.btn_export.pack_forget()

        self._sync_novo_btn()

    def _on_panel_state(self, is_editing: bool):
        self.btn_novo.configure(state="disabled" if is_editing else "normal")
        self.btn_export.configure(state="disabled" if is_editing else "normal")

    def _sync_novo_btn(self):
        panel   = self._active_panel()
        editing = panel.is_editing() if panel else False
        self.btn_novo.configure(state="disabled" if editing else "normal")

    def _active_panel(self):
        return (self._panel_insumos
                if self._aba_ativa == "Insumos"
                else self._panel_historico)

    def _on_novo(self):
        p = self._active_panel()
        if p: p.go_new()

    def _on_posicao_estoque(self):
        insumos = InsumoService().listar()
        if not insumos:
            messagebox.showinfo("Posição de Estoque", "Não há insumos para exportar."); return
            
        arq = filedialog.asksaveasfilename(
            title="Salvar posição de estoque", defaultextension=".xlsx",
            filetypes=[("Planilha Excel", "*.xlsx")],
            initialfile=f"posicao_estoque_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        )
        if not arq: return
        
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Estoque Atual"
            hdr = ["Nome", "Categoria", "Qtd Disponível", "Unidade", "Qtd Mínima", "Status", "Custo Un.", "Valor em Estoque"]
            ws.append(hdr)
            
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C96B7A")
                c.alignment = Alignment(horizontal="center", vertical="center")
                
            for i in insumos:
                status = "OK"
                if i.quantidade_disponivel <= 0: status = "CRÍTICO"
                elif i.quantidade_disponivel <= i.quantidade_minima: status = "ALERTA"
                
                valor_estoque = i.quantidade_disponivel * i.custo_por_unidade
                
                ws.append([
                    i.nome, i.categoria, i.quantidade_disponivel, i.unidade_medida,
                    i.quantidade_minima, status, i.custo_por_unidade, valor_estoque
                ])
                
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18
            ws.freeze_panes = "A2"; wb.save(arq)
            messagebox.showinfo("Sucesso", f"Posição de estoque exportada para:\n{arq}")
        except Exception as e:
            messagebox.showerror("Erro na Exportação", str(e))

    def _on_export(self):
        p = self._active_panel()
        if p: p.export_excel()
