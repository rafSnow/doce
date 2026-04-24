"""
FinanceiroView — Despesas e Rendimentos em uma única view com abas.
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.formatters import fmt_moeda, parse_data, parse_float
from app.models.despesa import Despesa
from app.models.rendimento import Rendimento
from app.services.despesa_service import DespesaService
from app.services.cliente_service import ClienteService
from app.services.rendimento_service import RendimentoService
from app.core.enums import StatusPagamento, CategoriaDespesa, FormaPagamento
from app.ui.theme import (
    ACCENT,
    BG_DEEP,
    CARD_BG,
    CARD_BORDER,
    COLOR_BLUE,
    COLOR_GREEN,
    COLOR_ORANGE,
    COLOR_PURPLE,
    FIELD_BG,
    HEADER_BG,
    TEXT_MUTED,
    TEXT_PRIMARY,
    TEXT_SECONDARY,
    _btn_accent,
    _btn_danger,
    _btn_ghost,
    _card,
    _combo,
    _entry,
    _optmenu,
    _sep,
    _treeview_style,
)


def _mask_date(event):
    e = event.widget
    d = "".join(c for c in e.get() if c.isdigit())[:8]
    parts = []
    if len(d) >= 2:  parts.append(d[:2])
    elif d:           parts.append(d)
    if len(d) >= 4:  parts.append(d[2:4])
    elif len(d) > 2: parts.append(d[2:])
    if len(d) > 4:   parts.append(d[4:])
    e.delete(0, "end"); e.insert(0, "/".join(parts))

def _mask_money(event):
    e = event.widget; raw = e.get()
    digits = "".join(c for c in raw if c.isdigit())
    if not digits:
        e.delete(0, "end"); return
    if "," in raw:
        pi, pd = raw.split(",", 1)
        int_d = "".join(c for c in pi if c.isdigit()) or "0"
        dec_d = "".join(c for c in pd if c.isdigit())
        if len(dec_d) > 2: int_d += dec_d[:-2]; dec_d = dec_d[-2:]
        else: dec_d = dec_d.ljust(2, "0")
    else:
        int_d = digits; dec_d = "00"
    groups = []; n = str(int(int_d))
    while n: groups.insert(0, n[-3:]); n = n[:-3]
    e.delete(0, "end"); e.insert(0, f"{'.'.join(groups)},{dec_d}")


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  PAINEL DE DESPESAS                                                         ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class DespesasPanel(ctk.CTkFrame):

    def __init__(self, master, on_state_change=None, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.service    = DespesaService()
        self.current_id = None
        self._on_state  = on_state_change
        self._editing   = False

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_lista()
        self._build_form()
        self._show_lista()
        self._load()

    def is_editing(self) -> bool: return self._editing
    def go_new(self):    self._reset_form(); self._show_form(editing=False)
    def export_excel(self): self._do_export()
    
    def refresh(self):
        """Atualiza a lista de despesas."""
        self._load_safe()

    # ── lista ─────────────────────────────────────────────────────────────
    def _build_lista(self):
        self._frame_lista = ctk.CTkFrame(self, fg_color="transparent")
        self._frame_lista.grid_columnconfigure(0, weight=1)
        self._frame_lista.grid_rowconfigure(1, weight=1)

        fc = _card(self._frame_lista)
        fc.grid(row=0, column=0, padx=4, pady=(0, 10), sticky="ew")
        fc.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fc, text="FILTROS", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, padx=14, pady=(10, 4), sticky="w")

        ctrl = ctk.CTkFrame(fc, fg_color="transparent")
        ctrl.grid(row=1, column=0, padx=14, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(ctrl, text="Período:", text_color=TEXT_SECONDARY).pack(side="left", padx=(0, 4))
        self.f_ini = _entry(ctrl, width=110, placeholder_text="DD/MM/AAAA")
        self.f_ini.pack(side="left", padx=(0, 4))
        self.f_ini.bind("<KeyRelease>", _mask_date)
        ctk.CTkLabel(ctrl, text="até", text_color=TEXT_SECONDARY).pack(side="left", padx=4)
        self.f_fim = _entry(ctrl, width=110, placeholder_text="DD/MM/AAAA")
        self.f_fim.pack(side="left", padx=(0, 12))
        self.f_fim.bind("<KeyRelease>", _mask_date)

        ctk.CTkLabel(ctrl, text="Categoria:", text_color=TEXT_SECONDARY).pack(side="left", padx=(0, 4))
        self.f_cat = _optmenu(ctrl, ["Todos", CategoriaDespesa.INSUMOS.value, CategoriaDespesa.INVESTIMENTOS.value, CategoriaDespesa.OUTROS.value])
        self.f_cat.pack(side="left", padx=(0, 12)); self.f_cat.set("Todos")

        ctk.CTkLabel(ctrl, text="Status:", text_color=TEXT_SECONDARY).pack(side="left", padx=(0, 4))
        self.f_status = _optmenu(ctrl, ["Todos", StatusPagamento.PENDENTE.value, StatusPagamento.PAGO.value])
        self.f_status.pack(side="left", padx=(0, 12)); self.f_status.set("Todos")

        _btn_accent(ctrl, "Buscar", self._load_safe).pack(side="left", padx=(0, 8))
        _btn_ghost(ctrl, "Limpar", self._clear_filters).pack(side="left")

        tf = _card(self._frame_lista)
        tf.grid(row=1, column=0, sticky="nsew", padx=4)
        tf.grid_columnconfigure(0, weight=1); tf.grid_rowconfigure(0, weight=1)

        cols  = ("id", "data", "categoria", "valor", "status", "responsavel")
        style = _treeview_style("Desp")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", style=style)
        for col, txt, w, anc in [
            ("id",          "ID",          50,  "center"),
            ("data",        "Data",        100, "center"),
            ("categoria",   "Categoria",   130, "center"),
            ("valor",       "Valor",       110, "e"),
            ("status",      "Status",      100, "center"),
            ("responsavel", "Responsável", 160, "w"),
        ]:
            self.tree.heading(col, text=txt); self.tree.column(col, width=w, anchor=anc)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        sb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        sb.grid(row=0, column=1, sticky="ns", pady=2)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.bind("<Double-1>", self._on_edit_selected)

        # totalizadores
        tc = _card(self._frame_lista)
        tc.grid(row=2, column=0, sticky="ew", padx=4, pady=(10, 0))
        tc.grid_columnconfigure((0, 1, 2, 3), weight=1)

        ctk.CTkLabel(tc, text="RESUMO", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, columnspan=4, padx=14, pady=(10, 8), sticky="w")

        rc = ctk.CTkFrame(tc, fg_color="transparent")
        rc.grid(row=1, column=0, columnspan=4, sticky="ew", padx=6, pady=(0, 10))
        rc.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self._lbl_insumos       = self._mini_card(rc, 0, "Insumos",       COLOR_GREEN)
        self._lbl_investimentos = self._mini_card(rc, 1, "Investimentos",  COLOR_BLUE)
        self._lbl_outros        = self._mini_card(rc, 2, "Outros",         COLOR_ORANGE)
        self._lbl_geral         = self._mini_card(rc, 3, "Total Geral",    COLOR_PURPLE)

    @staticmethod
    def _mini_card(parent, col, title, color):
        c = ctk.CTkFrame(parent, fg_color=HEADER_BG, corner_radius=10,
                         border_width=1, border_color=CARD_BORDER)
        c.grid(row=0, column=col, padx=5, sticky="ew")
        c.grid_columnconfigure(0, weight=1)
        ctk.CTkFrame(c, height=4, fg_color=color, corner_radius=2).grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(c, text=title.upper(), text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=9, weight="bold")
                     ).grid(row=1, column=0, padx=12, pady=(8, 0), sticky="w")
        lbl = ctk.CTkLabel(c, text="R$ 0,00", text_color=color,
                           font=ctk.CTkFont(family="Roboto", size=18, weight="bold"))
        lbl.grid(row=2, column=0, padx=12, pady=(2, 10), sticky="w")
        return lbl

    # ── formulário ────────────────────────────────────────────────────────
    def _build_form(self):
        self._frame_form = ctk.CTkFrame(self, fg_color="transparent")
        self._frame_form.grid_columnconfigure((0, 1), weight=1)

        # Atalhos de teclado
        self._frame_form.bind("<Control-s>", lambda e: self._on_save())
        self._frame_form.bind("<Escape>",    lambda e: self._show_lista())

        cg = _card(self._frame_form)
        cg.grid(row=0, column=0, columnspan=2, padx=4, pady=(0, 10), sticky="ew")
        cg.grid_columnconfigure((1, 3), weight=1)

        self.lbl_form_title = ctk.CTkLabel(cg, text="DADOS GERAIS", text_color=TEXT_MUTED,
                                           font=ctk.CTkFont(size=10, weight="bold")
                                           )
        self.lbl_form_title.grid(row=0, column=0, columnspan=4, padx=14, pady=(10, 8), sticky="w")

        for label, r, c, attr, ph, bind in [
            ("Data*",      0, 0, "entry_data",  "DD/MM/AAAA", _mask_date),
            ("Valor (R$)*",0, 2, "entry_valor", "0,00",       _mask_money),
        ]:
            ctk.CTkLabel(cg, text=label, text_color=TEXT_SECONDARY
                         ).grid(row=r+1, column=c, padx=10, pady=8, sticky="e")
            e = _entry(cg, placeholder_text=ph)
            e.grid(row=r+1, column=c+1, padx=10, pady=8, sticky="ew")
            if bind: e.bind("<KeyRelease>", bind)
            setattr(self, attr, e)

        ctk.CTkLabel(cg, text="Categoria*", text_color=TEXT_SECONDARY
                     ).grid(row=2, column=0, padx=10, pady=8, sticky="e")
        self.cb_cat = _optmenu(cg, [CategoriaDespesa.INSUMOS.value, CategoriaDespesa.INVESTIMENTOS.value, CategoriaDespesa.OUTROS.value])
        self.cb_cat.grid(row=2, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(cg, text="Status*", text_color=TEXT_SECONDARY
                     ).grid(row=2, column=2, padx=10, pady=8, sticky="e")
        self.cb_status_f = _optmenu(cg, [StatusPagamento.PENDENTE.value, StatusPagamento.PAGO.value])
        self.cb_status_f.grid(row=2, column=3, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(cg, text="Responsável", text_color=TEXT_SECONDARY
                     ).grid(row=3, column=0, padx=10, pady=8, sticky="e")
        self.entry_resp = _entry(cg, placeholder_text="Nome")
        self.entry_resp.grid(row=3, column=1, columnspan=3, padx=10, pady=8, sticky="ew")

        cp = _card(self._frame_form)
        cp.grid(row=1, column=0, columnspan=2, padx=4, pady=(0, 10), sticky="ew")
        cp.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(cp, text="PAGAMENTO", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, columnspan=4, padx=14, pady=(10, 8), sticky="w")

        ctk.CTkLabel(cp, text="Forma", text_color=TEXT_SECONDARY
                     ).grid(row=1, column=0, padx=10, pady=8, sticky="e")
        self.cb_forma = _optmenu(cp, [FormaPagamento.PIX.value, FormaPagamento.DINHEIRO.value, FormaPagamento.CARTAO_CREDITO.value,
                                      FormaPagamento.CARTAO_DEBITO.value, FormaPagamento.BOLETO.value, FormaPagamento.TRANSFERENCIA.value])
        self.cb_forma.grid(row=1, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(cp, text="Data Pagamento", text_color=TEXT_SECONDARY
                     ).grid(row=1, column=2, padx=10, pady=8, sticky="e")
        self.entry_dt_pag = _entry(cp, placeholder_text="DD/MM/AAAA")
        self.entry_dt_pag.grid(row=1, column=3, padx=10, pady=8, sticky="ew")
        self.entry_dt_pag.bind("<KeyRelease>", _mask_date)

        cd = _card(self._frame_form)
        cd.grid(row=2, column=0, columnspan=2, padx=4, pady=(0, 10), sticky="nsew")
        cd.grid_columnconfigure(0, weight=1)
        self._frame_form.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(cd, text="DESCRIÇÃO", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, padx=14, pady=(10, 8), sticky="w")
        self.txt_desc = ctk.CTkTextbox(cd, height=140, fg_color=FIELD_BG,
                                       border_color=CARD_BORDER, text_color=TEXT_PRIMARY)
        self.txt_desc.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="nsew")

        fa = ctk.CTkFrame(self._frame_form, fg_color="transparent")
        fa.grid(row=3, column=0, columnspan=2, padx=4, pady=(0, 4), sticky="e")
        self.btn_salvar  = _btn_accent(fa, "Salvar",   self._on_save)
        self.btn_excluir = _btn_danger(fa, "Excluir",  self._on_delete)
        self.btn_cancel  = _btn_ghost(fa,  "Cancelar", self._show_lista)
        for b in (self.btn_cancel, self.btn_excluir, self.btn_salvar):
            b.pack(side="right", padx=5)

        ctk.CTkLabel(self._frame_form, text="* Campos obrigatórios", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10)).grid(row=4, column=0, columnspan=2, padx=14, pady=(0, 5), sticky="w")

    # ── transições ────────────────────────────────────────────────────────
    def _show_lista(self):
        self._editing = False
        self._frame_form.grid_forget()
        self._frame_lista.grid(row=0, column=0, sticky="nsew")
        self._frame_lista.grid_columnconfigure(0, weight=1)
        self._frame_lista.grid_rowconfigure(1, weight=1)
        if self._on_state: self._on_state(False)

    def _show_form(self, editing: bool):
        self._editing = True
        self._frame_lista.grid_forget()
        self._frame_form.grid(row=0, column=0, sticky="nsew")
        self._frame_form.grid_columnconfigure((0, 1), weight=1)
        if self._on_state: self._on_state(True)
        self.entry_data.focus_set()

    # ── dados ─────────────────────────────────────────────────────────────
    def _load(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        di = parse_data(self.f_ini.get().strip(), "Data inicial")
        df = parse_data(self.f_fim.get().strip(), "Data final")
        if di and df and datetime.strptime(di, "%d/%m/%Y") > datetime.strptime(df, "%d/%m/%Y"):
            raise ValueError("A data inicial não pode ser maior que a data final.")
        cat = self.f_cat.get(); status = self.f_status.get()
        for d in self.service.listar(data_inicio=di, data_fim=df, categoria=cat, status=status):
            self.tree.insert("", "end", values=(
                d.id, d.data, d.categoria,
                f"R$ {fmt_moeda(d.valor)}", d.status, d.responsavel or "",
            ))
        totais = self.service.total_por_categoria(data_inicio=di, data_fim=df,
                                                   status=status, categoria=cat)
        ti = float(totais.get(CategoriaDespesa.INSUMOS.value, 0) or 0)
        tv = float(totais.get(CategoriaDespesa.INVESTIMENTOS.value, 0) or 0)
        to = float(totais.get(CategoriaDespesa.OUTROS.value, 0) or 0)
        self._lbl_insumos.configure(text=f"R$ {fmt_moeda(ti)}")
        self._lbl_investimentos.configure(text=f"R$ {fmt_moeda(tv)}")
        self._lbl_outros.configure(text=f"R$ {fmt_moeda(to)}")
        self._lbl_geral.configure(text=f"R$ {fmt_moeda(ti+tv+to)}")

    def _load_safe(self):
        try: self._load()
        except ValueError as e: messagebox.showerror("Erro", str(e))

    def _clear_filters(self):
        self.f_ini.delete(0, "end"); self.f_fim.delete(0, "end")
        self.f_cat.set("Todos"); self.f_status.set("Todos")
        self._load_safe()

    def _reset_form(self):
        self.current_id = None
        self.entry_data.delete(0, "end")
        self.entry_data.insert(0, datetime.now().strftime("%d/%m/%Y"))
        self.entry_valor.delete(0, "end"); self.entry_valor.insert(0, "0,00")
        self.cb_cat.set(CategoriaDespesa.OUTROS.value); self.cb_status_f.set(StatusPagamento.PENDENTE.value)
        self.cb_forma.set(FormaPagamento.PIX.value)
        self.entry_dt_pag.delete(0, "end"); self.entry_resp.delete(0, "end")
        self.txt_desc.delete("0.0", "end")
        self.btn_excluir.configure(state="disabled")

    def _on_edit_selected(self, _=None):
        sel = self.tree.selection()
        if not sel: return
        d = self.service.get_by_id(int(self.tree.item(sel[0])["values"][0]))
        if not d: return
        self._reset_form(); self.current_id = d.id
        self.entry_data.delete(0,"end"); self.entry_data.insert(0, d.data or "")
        self.entry_valor.delete(0,"end"); self.entry_valor.insert(0, fmt_moeda(d.valor or 0))
        self.cb_cat.set(d.categoria or CategoriaDespesa.OUTROS.value)
        self.cb_status_f.set(d.status or StatusPagamento.PENDENTE.value)
        self.cb_forma.set(d.forma_pagamento or FormaPagamento.PIX.value)
        self.entry_dt_pag.delete(0,"end"); self.entry_dt_pag.insert(0, d.data_pagamento_final or "")
        self.entry_resp.delete(0,"end"); self.entry_resp.insert(0, d.responsavel or "")
        self.txt_desc.delete("0.0","end"); self.txt_desc.insert("0.0", d.descricao or "")
        
        # ERP integration: block editing if originated from insumo purchase
        if d.origem:
            self.btn_salvar.configure(state="disabled")
            self.btn_excluir.configure(state="disabled")
            # Optionally show a message in the description or a label
            self.txt_desc.insert("1.0", ">>> REGISTRO AUTOMÁTICO (Origem: " + d.origem.upper() + ") <<<\n"
                                 "Edições e exclusões devem ser feitas no módulo de origem.\n\n")
        else:
            self.btn_salvar.configure(state="normal")
            self.btn_excluir.configure(state="normal")
            
        self._show_form(editing=True)

    def _on_save(self):
        try:
            data  = parse_data(self.entry_data.get().strip(), "Data", obrigatorio=True)
            valor = parse_float(self.entry_valor.get().strip(), "Valor", obrigatorio=True, minimo=0.01)
            cat   = self.cb_cat.get()
            if cat not in [CategoriaDespesa.INSUMOS.value, CategoriaDespesa.INVESTIMENTOS.value, CategoriaDespesa.OUTROS.value]: raise ValueError("Categoria inválida.")
            status = self.cb_status_f.get()
            if status not in [StatusPagamento.PENDENTE.value, StatusPagamento.PAGO.value]: raise ValueError("Status inválido.")
            dt_pag = parse_data(self.entry_dt_pag.get().strip(), "Data Pagamento")
            if status == StatusPagamento.PAGO.value and not dt_pag:
                raise ValueError("Para despesas pagas, informe a data do pagamento.")
            self.service.salvar(Despesa(
                id=self.current_id, data=data, valor=valor,
                descricao=self.txt_desc.get("0.0","end").strip() or None,
                categoria=cat, responsavel=self.entry_resp.get().strip() or None,
                status=status, forma_pagamento=self.cb_forma.get() or None,
                data_pagamento_final=dt_pag or None,
            ))
            self._load_safe(); self._show_lista()
        except ValueError as e: messagebox.showerror("Erro", str(e))

    def _on_delete(self):
        if not self.current_id: return
        desc = self.txt_desc.get("0.0","end").strip() or f"ID {self.current_id}"
        if messagebox.askyesno("Confirmar exclusão",
                               f"Excluir a despesa '{desc}'?\nEsta ação não pode ser desfeita."):
            self.service.excluir(self.current_id)
            self._load_safe(); self._show_lista()

    def _do_export(self):
        try:
            di = parse_data(self.f_ini.get().strip(), "Data inicial")
            df = parse_data(self.f_fim.get().strip(), "Data final")
            despesas = self.service.listar(data_inicio=di, data_fim=df,
                                           categoria=self.f_cat.get(), status=self.f_status.get())
            if not despesas:
                messagebox.showinfo("Exportar", "Sem despesas para exportar."); return
            arq = filedialog.asksaveasfilename(
                title="Salvar despesas", defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=f"despesas_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            )
            if not arq: return
            wb = Workbook(); ws = wb.active; ws.title = "Despesas"
            hdr = ["ID","Data","Categoria","Valor (R$)","Status","Forma Pag.","Responsável","Descrição"]
            ws.append(hdr)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C96B7A")
                c.alignment = Alignment(horizontal="center", vertical="center")
            for d in despesas:
                ws.append([d.id,d.data,d.categoria,d.valor,d.status,
                           d.forma_pagamento or "",d.responsavel or "",d.descricao or ""])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col)+2, 45)
            ws.freeze_panes = "A2"; wb.save(arq)
            messagebox.showinfo("Exportar", f"Salvo em:\n{arq}")
        except Exception as e: messagebox.showerror("Erro", str(e))


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  PAINEL DE RENDIMENTOS                                                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class RendimentosPanel(ctk.CTkFrame):

    def __init__(self, master, on_state_change=None, **kw):
        super().__init__(master, fg_color="transparent", **kw)
        self.service         = RendimentoService()
        self.cliente_service = ClienteService()
        self.current_id      = None
        self.cliente_map: dict[int, str] = {}
        self._on_state = on_state_change
        self._editing  = False

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_lista()
        self._build_form()
        self._show_lista()
        self._load_clients()
        self._load_safe()

    def is_editing(self) -> bool: return self._editing
    def go_new(self):
        if not self.cliente_map:
            messagebox.showwarning("Aviso","Cadastre ao menos um cliente antes de lançar rendimento.")
            return
        self._reset_form(); self._show_form(editing=False)
    def export_excel(self): self._do_export()
    def refresh_clients(self): self._load_clients()
    
    def refresh(self):
        """Atualiza clientes e a lista de rendimentos."""
        self._load_clients()
        self._load_safe()

    # ── lista ─────────────────────────────────────────────────────────────
    def _build_lista(self):
        self._frame_lista = ctk.CTkFrame(self, fg_color="transparent")
        self._frame_lista.grid_columnconfigure(0, weight=1)
        self._frame_lista.grid_rowconfigure(1, weight=1)

        fc = _card(self._frame_lista)
        fc.grid(row=0, column=0, padx=4, pady=(0, 10), sticky="ew")
        fc.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fc, text="FILTROS", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, padx=14, pady=(10, 4), sticky="w")

        ctrl = ctk.CTkFrame(fc, fg_color="transparent")
        ctrl.grid(row=1, column=0, padx=14, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(ctrl, text="Período:", text_color=TEXT_SECONDARY).pack(side="left", padx=(0, 4))
        self.f_ini = _entry(ctrl, width=110, placeholder_text="DD/MM/AAAA")
        self.f_ini.pack(side="left", padx=(0, 4))
        self.f_ini.bind("<KeyRelease>", _mask_date)
        ctk.CTkLabel(ctrl, text="até", text_color=TEXT_SECONDARY).pack(side="left", padx=4)
        self.f_fim = _entry(ctrl, width=110, placeholder_text="DD/MM/AAAA")
        self.f_fim.pack(side="left", padx=(0, 12))
        self.f_fim.bind("<KeyRelease>", _mask_date)

        ctk.CTkLabel(ctrl, text="Status:", text_color=TEXT_SECONDARY).pack(side="left", padx=(0, 4))
        self.f_status = _optmenu(ctrl, ["Todos", StatusPagamento.PENDENTE.value, StatusPagamento.RECEBIDO.value])
        self.f_status.pack(side="left", padx=(0, 12)); self.f_status.set("Todos")

        _btn_accent(ctrl, "Buscar", self._load_safe).pack(side="left", padx=(0, 8))
        _btn_ghost(ctrl, "Limpar", self._clear_filters).pack(side="left")

        tf = _card(self._frame_lista)
        tf.grid(row=1, column=0, sticky="nsew", padx=4)
        tf.grid_columnconfigure(0, weight=1); tf.grid_rowconfigure(0, weight=1)

        cols  = ("id","cliente","pag_ini","status_ini","pag_fin","status_fin","resp")
        style = _treeview_style("Rend")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", style=style)
        for col, txt, w, anc in [
            ("id",         "ID",           50,  "center"),
            ("cliente",    "Cliente",      200, "w"),
            ("pag_ini",    "Pag. Inicial", 120, "e"),
            ("status_ini", "Status Ini.",  110, "center"),
            ("pag_fin",    "Pag. Final",   120, "e"),
            ("status_fin", "Status Fin.",  110, "center"),
            ("resp",       "Responsável",  150, "w"),
        ]:
            self.tree.heading(col, text=txt); self.tree.column(col, width=w, anchor=anc)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        sb = ttk.Scrollbar(tf, orient="vertical", command=self.tree.yview)
        sb.grid(row=0, column=1, sticky="ns", pady=2)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.bind("<Double-1>", self._on_edit_selected)

    # ── formulário ────────────────────────────────────────────────────────
    def _build_form(self):
        self._frame_form = ctk.CTkFrame(self, fg_color="transparent")
        self._frame_form.grid_columnconfigure((0, 1), weight=1)

        # Atalhos de teclado
        self._frame_form.bind("<Control-s>", lambda e: self._on_save())
        self._frame_form.bind("<Escape>",    lambda e: self._show_lista())

        cf = _card(self._frame_form)
        cf.grid(row=0, column=0, columnspan=2, padx=4, pady=(0, 10), sticky="ew")
        cf.grid_columnconfigure((1, 3), weight=1)

        ctk.CTkLabel(cf, text="DADOS DO RENDIMENTO", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10, weight="bold")
                     ).grid(row=0, column=0, columnspan=4, padx=14, pady=(10, 8), sticky="w")

        ctk.CTkLabel(cf, text="Cliente*", text_color=TEXT_SECONDARY
                     ).grid(row=1, column=0, padx=10, pady=8, sticky="e")
        self.cb_cliente = _combo(cf, ["Carregando..."])
        self.cb_cliente.grid(row=1, column=1, padx=10, pady=8, sticky="ew")

        ctk.CTkLabel(cf, text="Responsável", text_color=TEXT_SECONDARY
                     ).grid(row=1, column=2, padx=10, pady=8, sticky="e")
        self.entry_resp = _entry(cf, placeholder_text="Nome")
        self.entry_resp.grid(row=1, column=3, padx=10, pady=8, sticky="ew")

        for r, prefix, attr_v, attr_d, attr_f, attr_s in [
            (2, "Valor Inicial (R$)", "entry_ini_v", "entry_ini_d", "cb_ini_f", "cb_ini_s"),
            (4, "Valor Final (R$)",   "entry_fin_v", "entry_fin_d", "cb_fin_f", "cb_fin_s"),
        ]:
            ctk.CTkLabel(cf, text=prefix, text_color=TEXT_SECONDARY
                         ).grid(row=r, column=0, padx=10, pady=8, sticky="e")
            ev = _entry(cf, placeholder_text="0,00")
            ev.grid(row=r, column=1, padx=10, pady=8, sticky="ew")
            ev.bind("<KeyRelease>", _mask_money); setattr(self, attr_v, ev)

            ctk.CTkLabel(cf, text="Data", text_color=TEXT_SECONDARY
                         ).grid(row=r, column=2, padx=10, pady=8, sticky="e")
            ed = _entry(cf, placeholder_text="DD/MM/AAAA")
            ed.grid(row=r, column=3, padx=10, pady=8, sticky="ew")
            ed.bind("<KeyRelease>", _mask_date); setattr(self, attr_d, ed)

            ctk.CTkLabel(cf, text="Forma", text_color=TEXT_SECONDARY
                         ).grid(row=r+1, column=0, padx=10, pady=8, sticky="e")
            ef = _combo(cf, [FormaPagamento.PIX.value, FormaPagamento.DINHEIRO.value, FormaPagamento.CARTAO_CREDITO.value,
                              FormaPagamento.CARTAO_DEBITO.value, FormaPagamento.TRANSFERENCIA.value])
            ef.grid(row=r+1, column=1, padx=10, pady=8, sticky="ew")
            setattr(self, attr_f, ef)

            ctk.CTkLabel(cf, text="Status", text_color=TEXT_SECONDARY
                         ).grid(row=r+1, column=2, padx=10, pady=8, sticky="e")
            es = _combo(cf, [StatusPagamento.PENDENTE.value, StatusPagamento.RECEBIDO.value])
            es.grid(row=r+1, column=3, padx=10, pady=8, sticky="ew")
            setattr(self, attr_s, es)

        fa = ctk.CTkFrame(self._frame_form, fg_color="transparent")
        fa.grid(row=1, column=0, columnspan=2, padx=4, pady=(0, 4), sticky="ew")
        
        # Elementos de link para pedido (Regra R-01)
        self.link_frame = ctk.CTkFrame(fa, fg_color="transparent")
        self.link_frame.pack(side="left", padx=10)
        
        self.lbl_link_pedido = ctk.CTkLabel(self.link_frame, text="", text_color=ACCENT, 
                                            font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_link_pedido.pack(side="left", padx=5)
        
        self.btn_ir_pedido = _btn_ghost(self.link_frame, "Ir ao Pedido", self._ir_ao_pedido, width=100)
        self.btn_ir_pedido.pack(side="left", padx=5)

        self.btn_salvar  = _btn_accent(fa, "Salvar",   self._on_save)
        self.btn_excluir = _btn_danger(fa, "Excluir",  self._on_delete)
        self.btn_cancel  = _btn_ghost(fa,  "Cancelar", self._show_lista)
        for b in (self.btn_cancel, self.btn_excluir, self.btn_salvar):
            b.pack(side="right", padx=5)

        ctk.CTkLabel(self._frame_form, text="* Campos obrigatórios", text_color=TEXT_MUTED,
                     font=ctk.CTkFont(size=10)).grid(row=2, column=0, columnspan=2, padx=14, pady=(0, 5), sticky="w")

    # ── transições ────────────────────────────────────────────────────────
    def _show_lista(self):
        self._editing = False
        self._frame_form.grid_forget()
        self._frame_lista.grid(row=0, column=0, sticky="nsew")
        self._frame_lista.grid_columnconfigure(0, weight=1)
        self._frame_lista.grid_rowconfigure(1, weight=1)
        if self._on_state: self._on_state(False)

    def _show_form(self, editing: bool):
        self._editing = True
        self._frame_lista.grid_forget()
        self._frame_form.grid(row=0, column=0, sticky="nsew")
        self._frame_form.grid_columnconfigure((0, 1), weight=1)
        if self._on_state: self._on_state(True)
        self.cb_cliente.focus_set()

    # ── dados ─────────────────────────────────────────────────────────────
    def _set_form_readonly(self, readonly: bool, pedido_id: int | None = None):
        st = "disabled" if readonly else "normal"
        self.cb_cliente.configure(state=st)
        self.entry_resp.configure(state=st)
        self.entry_ini_v.configure(state=st)
        self.entry_ini_d.configure(state=st)
        self.cb_ini_f.configure(state=st)
        self.cb_ini_s.configure(state=st)
        self.entry_fin_v.configure(state=st)
        self.entry_fin_d.configure(state=st)
        self.cb_fin_f.configure(state=st)
        self.cb_fin_s.configure(state=st)
        
        if readonly and pedido_id:
            self.lbl_link_pedido.configure(text=f"Para editar, acesse o Pedido #{pedido_id}.")
            self.link_frame.pack(side="left", padx=10)
        else:
            self.link_frame.pack_forget()

    def _ir_ao_pedido(self):
        """Regra R-01: Navega para a view de pedidos e abre o pedido atual."""
        if not self.current_pedido_id_vinc: return
        
        # Obtém referência para a MainWindow através da hierarquia de widgets
        try:
            main_window = self.winfo_toplevel()
            if hasattr(main_window, "show_pedidos_view"):
                main_window.show_pedidos_view(pedido_id=self.current_pedido_id_vinc)
        except Exception as e:
            messagebox.showerror("Erro de Navegação", f"Não foi possível abrir o pedido: {e}")

    def _load_clients(self):
        self.cliente_map.clear(); labels = []
        for c in self.cliente_service.listar():
            if c.id is None: continue
            lbl = f"{c.id} - {c.nome}"
            self.cliente_map[int(c.id)] = lbl; labels.append(lbl)
        if not labels: labels = ["Nenhum cliente cadastrado"]
        self.cb_cliente.configure(values=labels); self.cb_cliente.set(labels[0])

    def _load(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        di = parse_data(self.f_ini.get().strip(), "Data inicial")
        df = parse_data(self.f_fim.get().strip(), "Data final")
        for r in self.service.listar(data_inicio=di, data_fim=df, status=self.f_status.get()):
            nome = self.cliente_map.get(r.cliente_id, "—")
            self.tree.insert("", "end", values=(
                r.id, nome,
                f"R$ {fmt_moeda(r.pag_inicial_valor or 0)}", r.pag_inicial_status,
                f"R$ {fmt_moeda(r.pag_final_valor or 0)}",   r.pag_final_status,
                r.responsavel or "",
            ))

    def _load_safe(self):
        try: self._load()
        except ValueError as e: messagebox.showerror("Erro", str(e))

    def _clear_filters(self):
        self.f_ini.delete(0,"end"); self.f_fim.delete(0,"end")
        self.f_status.set("Todos"); self._load_safe()

    def _reset_form(self):
        self.current_id = None
        self.current_pedido_id_vinc = None
        self._set_form_readonly(False)
        vals = self.cb_cliente.cget("values")
        if vals: self.cb_cliente.set(vals[0])
        self.entry_resp.delete(0,"end")
        for attr in ("entry_ini_v","entry_fin_v"):
            e = getattr(self, attr); e.delete(0,"end"); e.insert(0,"0,00")
        for attr in ("entry_ini_d","entry_fin_d"):
            getattr(self, attr).delete(0,"end")
        for attr in ("cb_ini_f","cb_fin_f"):
            getattr(self, attr).set(FormaPagamento.PIX.value)
        for attr in ("cb_ini_s","cb_fin_s"):
            getattr(self, attr).set(StatusPagamento.PENDENTE.value)
        self.btn_excluir.configure(state="disabled")

    def _on_edit_selected(self, _=None):
        sel = self.tree.selection()
        if not sel: return
        r = self.service.get_by_id(int(self.tree.item(sel[0])["values"][0]))
        if not r: return
        self._reset_form()
        self.current_id = r.id
        self.current_pedido_id_vinc = r.pedido_id
        
        if r.cliente_id in self.cliente_map:
            self.cb_cliente.set(self.cliente_map[r.cliente_id])
        self.entry_resp.delete(0,"end"); self.entry_resp.insert(0, r.responsavel or "")
        self.entry_ini_v.delete(0,"end"); self.entry_ini_v.insert(0, fmt_moeda(r.pag_inicial_valor or 0))
        self.entry_ini_d.delete(0,"end"); self.entry_ini_d.insert(0, r.pag_inicial_data or "")
        self.cb_ini_f.set(r.pag_inicial_forma or FormaPagamento.PIX.value)
        self.cb_ini_s.set(r.pag_inicial_status or StatusPagamento.PENDENTE.value)
        self.entry_fin_v.delete(0,"end"); self.entry_fin_v.insert(0, fmt_moeda(r.pag_final_valor or 0))
        self.entry_fin_d.delete(0,"end"); self.entry_fin_d.insert(0, r.pag_final_data or "")
        self.cb_fin_f.set(r.pag_final_forma or FormaPagamento.PIX.value)
        self.cb_fin_s.set(r.pag_final_status or StatusPagamento.PENDENTE.value)
        
        # ERP integration: block editing if originated from a Pedido (Rule R-01)
        if r.pedido_id:
            self._set_form_readonly(True, pedido_id=r.pedido_id)
            self.btn_salvar.configure(state="disabled")
            self.btn_excluir.configure(state="disabled")
        else:
            self.btn_salvar.configure(state="normal")
            self.btn_excluir.configure(state="normal")
            
        self._show_form(editing=True)

    def _on_save(self):
        try:
            lbl = self.cb_cliente.get().strip()
            if " - " not in lbl: raise ValueError("Selecione um cliente válido.")
            cid   = int(lbl.split(" - ", 1)[0])
            ini_v = parse_float(self.entry_ini_v.get().strip(), "Valor Inicial", minimo=0)
            ini_d = parse_data(self.entry_ini_d.get().strip(), "Data Inicial")
            ini_f = self.cb_ini_f.get().strip() or None
            ini_s = self.cb_ini_s.get().strip()
            fin_v = parse_float(self.entry_fin_v.get().strip(), "Valor Final", minimo=0)
            fin_d = parse_data(self.entry_fin_d.get().strip(), "Data Final")
            fin_f = self.cb_fin_f.get().strip() or None
            fin_s = self.cb_fin_s.get().strip()
            if ini_s == StatusPagamento.RECEBIDO.value and not ini_d:
                raise ValueError("Informe a data do pagamento inicial recebido.")
            if fin_s == StatusPagamento.RECEBIDO.value and not fin_d:
                raise ValueError("Informe a data do pagamento final recebido.")
            self.service.salvar(Rendimento(
                id=self.current_id, cliente_id=cid,
                pag_inicial_valor=ini_v, pag_inicial_data=ini_d or None,
                pag_inicial_forma=ini_f, pag_inicial_status=ini_s,
                pag_final_valor=fin_v,   pag_final_data=fin_d or None,
                pag_final_forma=fin_f,   pag_final_status=fin_s,
                responsavel=self.entry_resp.get().strip() or None,
            ))
            self._load_safe(); self._show_lista()
        except ValueError as e: messagebox.showerror("Erro", str(e))

    def _on_delete(self):
        if not self.current_id: return
        if messagebox.askyesno("Confirmar exclusão",
                               f"Excluir rendimento ID {self.current_id}?\nEsta ação não pode ser desfeita."):
            self.service.excluir(self.current_id)
            self._load_safe(); self._show_lista()

    def _do_export(self):
        try:
            di = parse_data(self.f_ini.get().strip(), "Data inicial")
            df = parse_data(self.f_fim.get().strip(), "Data final")
            rends = self.service.listar(data_inicio=di, data_fim=df, status=self.f_status.get())
            if not rends:
                messagebox.showinfo("Exportar","Sem rendimentos para exportar."); return
            arq = filedialog.asksaveasfilename(
                title="Salvar rendimentos", defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],
                initialfile=f"rendimentos_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            )
            if not arq: return
            wb = Workbook(); ws = wb.active; ws.title = "Rendimentos"
            hdr = ["ID","Cliente","Valor Ini.","Data Ini.","Forma Ini.","Status Ini.",
                   "Valor Fin.","Data Fin.","Forma Fin.","Status Fin.","Responsável"]
            ws.append(hdr)
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="C96B7A")
                c.alignment = Alignment(horizontal="center", vertical="center")
            for r in rends:
                ws.append([r.id, self.cliente_map.get(r.cliente_id,"—"),
                           r.pag_inicial_valor, r.pag_inicial_data, r.pag_inicial_forma, r.pag_inicial_status,
                           r.pag_final_valor, r.pag_final_data, r.pag_final_forma, r.pag_final_status,
                           r.responsavel or ""])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 2, 45)
            ws.freeze_panes = "A2"; wb.save(arq)
            messagebox.showinfo("Exportar", f"Salvo em:\n{arq}")
        except Exception as e: messagebox.showerror("Erro", str(e))


# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║  FINANCEIRO VIEW                                                            ║
# ╚══════════════════════════════════════════════════════════════════════════════╝
class FinanceiroView(ctk.CTkFrame):

    ABAS = ["Despesas", "Rendimentos"]

    def __init__(self, master, **kw):
        super().__init__(master, fg_color=BG_DEEP, **kw)
        self._aba_ativa = "Despesas"
        self._tab_btns: dict[str, tk.Frame]     = {}
        self._tab_lbls: dict[str, ctk.CTkLabel] = {}

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self._build_topbar()
        self._build_tabbar()
        self._build_content()
        self._switch_aba("Despesas")

    def refresh(self):
        self._panel_despesas._load()
        self._panel_rendimentos.refresh_clients()
        self._panel_rendimentos._load()

    def _build_topbar(self):
        bar = ctk.CTkFrame(self, fg_color=HEADER_BG, corner_radius=0, height=56)
        bar.grid(row=0, column=0, sticky="ew")
        bar.grid_propagate(False)
        bar.grid_columnconfigure(1, weight=1)

        self.lbl_title = ctk.CTkLabel(
            bar, text="Financeiro — Despesas",
            text_color=TEXT_PRIMARY,
            font=ctk.CTkFont(family="Roboto", size=20, weight="bold"),
        )
        self.lbl_title.grid(row=0, column=0, padx=20, sticky="w")

        self._action_bar = ctk.CTkFrame(bar, fg_color="transparent")
        self._action_bar.grid(row=0, column=1, padx=(6, 16), sticky="e")

        self.btn_novo   = _btn_accent(self._action_bar, "+ Novo",        self._on_novo)
        self.btn_novo.pack(side="left", padx=4)

        self.btn_export = _btn_ghost(self._action_bar,  "Exportar Excel", self._on_export)
        self.btn_export.pack(side="left", padx=4)

    def _build_tabbar(self):
        self._tabbar = ctk.CTkFrame(self, fg_color=HEADER_BG, corner_radius=0, height=40)
        self._tabbar.grid(row=1, column=0, sticky="ew")
        self._tabbar.grid_propagate(False)

        for aba in self.ABAS:
            tab = tk.Frame(self._tabbar, bg=HEADER_BG, cursor="hand2")
            tab.pack(side="left", padx=(16, 0))
            self._tab_btns[aba] = tab

            lbl = ctk.CTkLabel(tab, text=aba, fg_color="transparent",
                               text_color=TEXT_MUTED, font=ctk.CTkFont(size=13))
            lbl.pack(pady=(8, 0))
            self._tab_lbls[aba] = lbl

            ind = tk.Frame(tab, bg=HEADER_BG, height=2)
            ind.pack(fill="x", pady=(4, 0))
            tab._indicator = ind  # type: ignore

            tab.bind("<Button-1>", lambda e, a=aba: self._switch_aba(a))
            lbl.bind("<Button-1>",  lambda e, a=aba: self._switch_aba(a))

        _sep(self).grid(row=1, column=0, sticky="sew")

    def _build_content(self):
        self._content = ctk.CTkFrame(self, fg_color="transparent")
        self._content.grid(row=2, column=0, sticky="nsew", padx=14, pady=(10, 14))
        self._content.grid_columnconfigure(0, weight=1)
        self._content.grid_rowconfigure(0, weight=1)

        self._panel_despesas    = DespesasPanel(self._content, on_state_change=self._on_panel_state)
        self._panel_rendimentos = RendimentosPanel(self._content, on_state_change=self._on_panel_state)

    def _switch_aba(self, aba: str):
        self._aba_ativa = aba
        for nome, tab in self._tab_btns.items():
            ativo = nome == aba
            self._tab_lbls[nome].configure(
                text_color=TEXT_PRIMARY if ativo else TEXT_MUTED,
                font=ctk.CTkFont(size=13, weight="bold" if ativo else "normal"),
            )
            tab._indicator.configure(bg=ACCENT if ativo else HEADER_BG)

        self.lbl_title.configure(text=f"Financeiro — {aba}")
        self._panel_despesas.grid_forget()
        self._panel_rendimentos.grid_forget()

        panel = self._active_panel()
        if panel:
            panel.grid(row=0, column=0, sticky="nsew")
            if hasattr(panel, "refresh"):
                panel.refresh()

        self._sync_novo_btn()

    def _on_panel_state(self, is_editing: bool):
        self.btn_novo.configure(state="disabled" if is_editing else "normal")
        self.btn_export.configure(state="disabled" if is_editing else "normal")

    def _sync_novo_btn(self):
        panel = self._active_panel()
        self.btn_novo.configure(state="disabled" if (panel and panel.is_editing()) else "normal")

    def _active_panel(self):
        return self._panel_despesas if self._aba_ativa == "Despesas" else self._panel_rendimentos

    def _on_novo(self):
        p = self._active_panel()
        if p: p.go_new()

    def _on_export(self):
        p = self._active_panel()
        if p: p.export_excel()
